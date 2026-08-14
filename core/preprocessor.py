"""Spreadsheet preprocessing, profiling, and reusable local caching."""

from __future__ import annotations

import hashlib
import json
import logging
import math
import random
import re
import shutil
import zipfile
from collections import Counter, deque
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from collections.abc import Callable
from typing import Any

import duckdb
import pandas as pd
import psutil
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries

from config.settings import settings


logger = logging.getLogger(__name__)


_DIMENSION_PATTERN = re.compile(
    rb"<(?:[A-Za-z_][\w.-]*:)?dimension\b[^>]*\bref\s*=\s*(['\"])([^'\"]+)\1",
    re.IGNORECASE,
)
_ROW_BEYOND_FIRST_PATTERN = re.compile(rb"<row\b[^>]*\br\s*=\s*['\"](?:[2-9]|[1-9][0-9]+)['\"]")
_CELL_BEYOND_A1_PATTERN = re.compile(
    rb"<c\b[^>]*\br\s*=\s*['\"](?:[B-Z]|[A-Z]{2,})[1-9][0-9]*['\"]",
    re.IGNORECASE,
)


class ProcessingCancelled(Exception):
    """Raised when the user cancels an in-flight import."""


class SchemaContamination(Exception):
    """Raised when a later streamed value does not fit the inferred schema."""

    def __init__(self, column: str, value: Any) -> None:
        self.column = column
        self.value = value
        super().__init__(f"Column {column!r} contains a conflicting value: {value!r}")


@dataclass(frozen=True)
class XlsxWorksheetInfo:
    """Cheap ZIP-level worksheet facts used for import routing and safeguards."""

    archive_path: str
    compressed_bytes: int
    uncompressed_bytes: int
    declared_dimension: str | None = None
    declared_max_row: int = 0
    declared_max_column: int = 0
    has_data_beyond_a1: bool = False


@dataclass(frozen=True)
class XlsxImportPlan:
    """Backend-only import route selected without changing the UI contract."""

    mode: str
    source_bytes: int
    uncompressed_bytes: int
    compression_ratio: float
    shared_strings_bytes: int
    largest_worksheet_bytes: int
    worksheets: tuple[XlsxWorksheetInfo, ...] = ()
    warnings: tuple[str, ...] = ()

    def worksheet(self, archive_path: str) -> XlsxWorksheetInfo | None:
        normalized = str(archive_path).replace("\\", "/").lstrip("/")
        for item in self.worksheets:
            if item.archive_path == normalized:
                return item
        return None


@dataclass
class SheetMeta:
    sheet_name: str
    rows: int
    cols: int
    columns: list[str]
    dtypes: dict[str, str]
    null_counts: dict[str, int]
    head_sample: list[dict[str, Any]]
    describe: dict[str, Any]
    unique_values: dict[str, list[str]]
    sheet_id: str = ""
    cache_path: str = ""
    sample_cache_path: str = ""
    unique_counts: dict[str, int] = field(default_factory=dict)
    unique_rates: dict[str, float] = field(default_factory=dict)
    type_profiles: dict[str, dict[str, Any]] = field(default_factory=dict)
    semantic_roles: dict[str, list[str]] = field(default_factory=dict)
    header_mode: str = "own"
    header_source_sheet_id: str = ""
    first_row_is_data: bool = False
    continuation_detected: bool = False
    continuation_confidence: float = 0.0
    continuation_reasons: list[str] = field(default_factory=list)

    def to_prompt_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet_name,
            "sheet_id": self.sheet_id,
            "shape": [self.rows, self.cols],
            "columns": self.columns,
            "dtypes": self.dtypes,
            "null_counts": self.null_counts,
            "unique_counts": self.unique_counts,
            "unique_rates": self.unique_rates,
            "type_profiles": self.type_profiles,
            "sample": self.head_sample,
            "describe": self.describe,
            "unique_values": self.unique_values,
            "semantic_roles": self.semantic_roles,
            "header_detection": {
                "mode": self.header_mode,
                "source_sheet_id": self.header_source_sheet_id,
                "first_row_is_data": self.first_row_is_data,
                "continuation_detected": self.continuation_detected,
                "confidence": self.continuation_confidence,
                "reasons": self.continuation_reasons,
            },
        }


@dataclass
class SheetGroupMeta:
    group_id: str
    group_type: str
    sheet_ids: list[str]
    sheet_names: list[str]
    columns: list[str]
    dtype_families: dict[str, str]
    total_rows: int
    confidence: float
    reason: str

    def to_prompt_dict(self) -> dict[str, Any]:
        return {
            "group_id": self.group_id,
            "type": self.group_type,
            "sheet_ids": self.sheet_ids,
            "sheet_names": self.sheet_names,
            "columns": self.columns,
            "dtype_families": self.dtype_families,
            "total_rows": self.total_rows,
            "confidence": self.confidence,
            "reason": self.reason,
            "runtime_access": (
                f'data.union_sheets("<dataset_id>", group_id="{self.group_id}", '
                "columns=[...])"
            ),
        }


@dataclass
class FileMeta:
    file_path: str
    file_name: str
    file_size_kb: float
    sheet_count: int
    sheets: list[SheetMeta] = field(default_factory=list)
    sheet_groups: list[SheetGroupMeta] = field(default_factory=list)
    dataset_id: str = ""
    display_name: str = ""
    source_fingerprint: str = ""
    content_hash: str = ""
    schema_family_id: str = ""
    profile_mode: str = "full"
    profile_sample_rows: int = 0

    @property
    def runtime_key(self) -> str:
        return self.dataset_id or self.file_name

    def to_prompt_dict(self) -> dict[str, Any]:
        dataset_key = self.runtime_key
        return {
            "file": self.file_name,
            "display_name": self.display_name or self.file_name,
            "dataset_id": dataset_key,
            "runtime_access": {
                "preferred": f'data.get("{dataset_key}", "<sheet_id>")',
                "compatible": f'dfs["{dataset_key}"]["<sheet_id>"]',
            },
            "size_kb": round(self.file_size_kb, 1),
            "profile_mode": self.profile_mode,
            "profile_sample_rows": self.profile_sample_rows,
            "sheet_groups": [
                group.to_prompt_dict()
                for group in self.sheet_groups
            ],
            "sheets": [sheet.to_prompt_dict() for sheet in self.sheets],
        }


class Preprocessor:
    """Read workbooks, build caches, and prepare profile metadata."""

    def __init__(self) -> None:
        self._preview_rows = settings.PREVIEW_ROWS
        self._max_cols_describe = settings.MAX_COLS_DESCRIBE
        self._max_unique_values = settings.MAX_PROFILE_UNIQUES
        self._sample_rows = settings.SAMPLE_ROWS_PER_SHEET
        self._batch_rows = settings.IMPORT_BATCH_ROWS
        self._schema_sample_rows = settings.IMPORT_SCHEMA_SAMPLE_ROWS
        self._row_group_size = settings.IMPORT_ROW_GROUP_SIZE
        self._batch_target_bytes = settings.IMPORT_BATCH_TARGET_BYTES
        self._batch_memory_fraction = settings.IMPORT_BATCH_MEMORY_FRACTION
        self._min_batch_rows = settings.IMPORT_MIN_BATCH_ROWS
        self._cancel_check_rows = settings.IMPORT_CANCEL_CHECK_ROWS

    def process(
        self,
        file_path: str,
        progress_callback: Callable[[dict[str, Any]], None] | None = None,
        cancel_callback: Callable[[], bool] | None = None,
    ) -> FileMeta:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File does not exist: {file_path}")
        if path.suffix.lower() not in {".xlsx", ".xls", ".xlsm"}:
            raise ValueError(f"Unsupported file type: {path.suffix}")

        size_bytes = path.stat().st_size
        if size_bytes > settings.MAX_DATASET_BYTES:
            raise ValueError("File exceeds the 1 GiB per-dataset limit.")

        self._raise_if_cancelled(cancel_callback)
        import_plan: XlsxImportPlan | None = None
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            import_plan = self._inspect_xlsx_package(path)
            self._validate_xlsx_import_plan(import_plan)

        settings.DATASET_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        settings.DUCKDB_TEMP_DIR.mkdir(parents=True, exist_ok=True)
        self._ensure_free_disk_space(path, import_plan=import_plan)

        self._emit_progress(progress_callback, "fingerprinting", 2, path.name)
        fingerprint = self._fingerprint(path)
        content_hash = self._content_hash(
            path,
            cancel_callback=cancel_callback,
        )
        dataset_id = f"ds_{fingerprint[:12]}"
        size_kb = size_bytes / 1024
        self._emit_progress(progress_callback, "inspecting", 5, path.name)
        logger.info(
            "Starting preprocessing for %s size=%.2fMB dataset_id=%s",
            path,
            size_bytes / (1024 * 1024),
            dataset_id,
        )

        if path.suffix.lower() == ".xls" and size_kb >= settings.LARGE_EXCEL_MB * 1024:
            raise ValueError(
                "Large legacy .xls files cannot be streamed safely. Save the workbook "
                "as .xlsx and import it again."
            )

        if path.suffix.lower() in {".xlsx", ".xlsm"} and self._should_stream_xlsx(
            path,
            size_kb,
            cancel_callback=cancel_callback,
            import_plan=import_plan,
        ):
            self._emit_progress(progress_callback, "importing", 10, path.name)
            sheets = self._process_large_xlsx(
                path,
                dataset_id,
                progress_callback=progress_callback,
                cancel_callback=cancel_callback,
                import_plan=import_plan,
            )
            profile_mode = "sampled"
        else:
            sheets = self._process_small_workbook(
                path,
                dataset_id,
                progress_callback=progress_callback,
                cancel_callback=cancel_callback,
            )
            profile_mode = "full"

        sheet_groups = self._detect_sheet_groups(dataset_id, sheets)
        self._emit_progress(progress_callback, "ready", 100, path.name)
        file_meta = FileMeta(
            file_path=str(path.resolve()),
            file_name=path.name,
            display_name=path.name,
            dataset_id=dataset_id,
            source_fingerprint=fingerprint,
            content_hash=content_hash,
            schema_family_id=self.schema_family_id(sheets),
            file_size_kb=size_kb,
            sheet_count=len(sheets),
            sheets=sheets,
            sheet_groups=sheet_groups,
            profile_mode=profile_mode,
            profile_sample_rows=(
                self._sample_rows if profile_mode == "sampled" else sum(sheet.rows for sheet in sheets)
            ),
        )
        self._write_dataset_manifest(dataset_id, file_meta)
        logger.info(
            "Finished preprocessing for %s dataset_id=%s profile_mode=%s sheets=%s",
            path.name,
            dataset_id,
            profile_mode,
            len(sheets),
        )
        return file_meta

    @classmethod
    def _inspect_xlsx_package(cls, path: Path) -> XlsxImportPlan:
        """Inspect ZIP metadata and a bounded XML prefix without inflating sheets."""
        try:
            with zipfile.ZipFile(path) as archive:
                entries = archive.infolist()
                uncompressed_bytes = sum(int(info.file_size) for info in entries)
                shared_strings_bytes = 0
                worksheet_infos: list[XlsxWorksheetInfo] = []
                probe_bytes = max(4096, int(settings.IMPORT_DIMENSION_PROBE_BYTES))

                for info in entries:
                    normalized = str(info.filename).replace("\\", "/").lstrip("/")
                    lowered = normalized.lower()
                    if lowered == "xl/sharedstrings.xml":
                        shared_strings_bytes = int(info.file_size)
                    if not (
                        lowered.startswith("xl/worksheets/")
                        and lowered.endswith(".xml")
                    ):
                        continue

                    with archive.open(info) as source:
                        prefix = source.read(probe_bytes)
                    dimension = None
                    declared_max_row = 0
                    declared_max_column = 0
                    match = _DIMENSION_PATTERN.search(prefix)
                    if match is not None:
                        dimension = match.group(2).decode("ascii", errors="ignore") or None
                    if dimension:
                        try:
                            _, _, declared_max_column, declared_max_row = range_boundaries(
                                dimension.replace("$", "")
                            )
                        except (TypeError, ValueError):
                            logger.warning(
                                "Invalid worksheet dimension path=%s dimension=%r",
                                normalized,
                                dimension,
                            )
                            declared_max_row = 0
                            declared_max_column = 0

                    has_data_beyond_a1 = bool(
                        _ROW_BEYOND_FIRST_PATTERN.search(prefix)
                        or _CELL_BEYOND_A1_PATTERN.search(prefix)
                    )
                    worksheet_infos.append(
                        XlsxWorksheetInfo(
                            archive_path=normalized,
                            compressed_bytes=int(info.compress_size),
                            uncompressed_bytes=int(info.file_size),
                            declared_dimension=dimension,
                            declared_max_row=declared_max_row,
                            declared_max_column=declared_max_column,
                            has_data_beyond_a1=has_data_beyond_a1,
                        )
                    )
        except zipfile.BadZipFile as exc:
            raise ValueError("Workbook is not a valid XLSX ZIP package.") from exc

        source_bytes = int(path.stat().st_size)
        compression_ratio = uncompressed_bytes / max(1, source_bytes)
        largest_worksheet_bytes = max(
            (item.uncompressed_bytes for item in worksheet_infos),
            default=0,
        )
        suspicious_dimensions = [
            item
            for item in worksheet_infos
            if (
                item.declared_dimension in {"A1", "A1:A1", None}
                and (
                    item.has_data_beyond_a1
                    or item.uncompressed_bytes
                    >= settings.IMPORT_SUSPICIOUS_DIMENSION_XML_BYTES
                )
            )
        ]
        declared_large = any(
            item.declared_max_row >= settings.LARGE_DATASET_ROWS
            for item in worksheet_infos
        )
        safe_mode = any(
            (
                source_bytes >= settings.IMPORT_SAFE_SOURCE_BYTES,
                uncompressed_bytes >= settings.IMPORT_SAFE_UNCOMPRESSED_BYTES,
                largest_worksheet_bytes >= settings.IMPORT_SAFE_WORKSHEET_XML_BYTES,
                shared_strings_bytes >= settings.IMPORT_SAFE_SHARED_STRINGS_BYTES,
            )
        )
        source_large = source_bytes >= max(0, settings.LARGE_EXCEL_MB) * 1024 * 1024
        fast_uncompressed = (
            uncompressed_bytes <= settings.IMPORT_FAST_UNCOMPRESSED_BYTES
        )
        if safe_mode:
            mode = "safe"
        elif source_large or declared_large or suspicious_dimensions or not fast_uncompressed:
            mode = "stream"
        else:
            mode = "small"

        warnings: list[str] = []
        for item in suspicious_dimensions:
            warnings.append(
                f"{item.archive_path}: declared dimension "
                f"{item.declared_dimension or '<missing>'} is not trusted"
            )
        return XlsxImportPlan(
            mode=mode,
            source_bytes=source_bytes,
            uncompressed_bytes=uncompressed_bytes,
            compression_ratio=round(compression_ratio, 4),
            shared_strings_bytes=shared_strings_bytes,
            largest_worksheet_bytes=largest_worksheet_bytes,
            worksheets=tuple(worksheet_infos),
            warnings=tuple(warnings),
        )

    @staticmethod
    def _validate_xlsx_import_plan(plan: XlsxImportPlan) -> None:
        if plan.uncompressed_bytes > settings.IMPORT_MAX_UNCOMPRESSED_BYTES:
            raise ValueError(
                "Workbook expands beyond the configured safe import limit "
                f"({plan.uncompressed_bytes / (1024**3):.1f} GiB)."
            )
        if (
            plan.uncompressed_bytes >= settings.IMPORT_ZIP_BOMB_MIN_UNCOMPRESSED_BYTES
            and plan.compression_ratio > settings.IMPORT_MAX_COMPRESSION_RATIO
        ):
            raise ValueError(
                "Workbook compression ratio is too high for safe import "
                f"({plan.compression_ratio:.1f}x)."
            )
        if plan.shared_strings_bytes > settings.IMPORT_MAX_SHARED_STRINGS_BYTES:
            raise ValueError(
                "Workbook shared strings exceed the configured safe import limit."
            )

        available_memory = max(1, int(psutil.virtual_memory().available))
        estimated_shared_memory = int(
            plan.shared_strings_bytes
            * settings.IMPORT_SHARED_STRINGS_MEMORY_MULTIPLIER
        )
        shared_memory_budget = int(
            available_memory
            * settings.IMPORT_SHARED_STRINGS_MAX_MEMORY_FRACTION
        )
        if estimated_shared_memory > shared_memory_budget:
            raise ValueError(
                "Workbook shared strings are too large for the currently available memory."
            )

        logger.info(
            "XLSX preflight mode=%s source=%.2fMB uncompressed=%.2fMB "
            "ratio=%.2fx shared_strings=%.2fMB worksheets=%s warnings=%s",
            plan.mode,
            plan.source_bytes / (1024 * 1024),
            plan.uncompressed_bytes / (1024 * 1024),
            plan.compression_ratio,
            plan.shared_strings_bytes / (1024 * 1024),
            len(plan.worksheets),
            list(plan.warnings),
        )

    @classmethod
    def _detect_sheet_groups(
        cls,
        dataset_id: str,
        sheets: list[SheetMeta],
    ) -> list[SheetGroupMeta]:
        buckets: dict[tuple[tuple[str, ...], tuple[str, ...]], list[SheetMeta]] = {}
        for sheet in sheets:
            if not sheet.columns:
                continue
            columns = tuple(cls._normalize_group_column(column) for column in sheet.columns)
            dtype_families = tuple(
                cls._dtype_family(sheet.dtypes.get(column, ""))
                for column in sheet.columns
            )
            buckets.setdefault((columns, dtype_families), []).append(sheet)

        groups: list[SheetGroupMeta] = []
        for (columns_signature, dtype_signature), grouped_sheets in buckets.items():
            if len(grouped_sheets) < 2:
                continue
            raw_columns = list(grouped_sheets[0].columns)
            sheet_ids = [
                sheet.sheet_id or sheet.sheet_name
                for sheet in grouped_sheets
            ]
            signature_payload = json.dumps(
                {
                    "dataset_id": dataset_id,
                    "columns": columns_signature,
                    "dtypes": dtype_signature,
                    "sheet_ids": sheet_ids,
                },
                ensure_ascii=False,
                separators=(",", ":"),
            )
            group_hash = hashlib.sha256(
                signature_payload.encode("utf-8")
            ).hexdigest()[:12]
            confidence = cls._sheet_group_confidence(grouped_sheets)
            groups.append(
                SheetGroupMeta(
                    group_id=f"sg_{group_hash}",
                    group_type="same_schema_append",
                    sheet_ids=sheet_ids,
                    sheet_names=[sheet.sheet_name for sheet in grouped_sheets],
                    columns=raw_columns,
                    dtype_families=dict(zip(raw_columns, dtype_signature)),
                    total_rows=sum(sheet.rows for sheet in grouped_sheets),
                    confidence=confidence,
                    reason=(
                        "Sheets share the same column order and compatible type "
                        "families, so they can be appended with UNION ALL."
                    ),
                )
            )
        return sorted(groups, key=lambda group: group.sheet_names)

    @staticmethod
    def _normalize_group_column(column: str) -> str:
        return str(column)

    @staticmethod
    def _dtype_family(dtype: str) -> str:
        normalized = str(dtype).lower()
        if any(token in normalized for token in ("int", "float", "double", "decimal")):
            return "number"
        if any(token in normalized for token in ("date", "time", "timestamp", "datetime")):
            return "datetime"
        if "bool" in normalized:
            return "boolean"
        if any(token in normalized for token in ("string", "object", "utf8", "large_string")):
            return "text"
        return normalized or "unknown"

    def _detect_headerless_continuation(
        self,
        previous_sheet: SheetMeta | None,
        first_row: Any,
        following_rows: list[Any],
    ) -> tuple[bool, list[str]]:
        """Conservatively identify a headerless continuation worksheet.

        Every rule below is mandatory.  Workbook size, worksheet names, row-key
        continuity, and other circumstantial signals intentionally have no role
        in the decision.
        """
        if previous_sheet is None or not previous_sheet.columns:
            return False, []

        expected_headers = list(previous_sheet.columns)
        expected_width = len(expected_headers)
        raw_rows = [first_row, *following_rows]
        populated_rows = [
            row
            for row in raw_rows
            if row is not None
            and any(value not in (None, "") for value in row)
        ]
        if len(populated_rows) < 3:
            return False, []

        for row in populated_rows:
            values = list(row)
            if len(values) < expected_width:
                return False, []
            overflow = values[expected_width:]
            if any(value not in (None, "") for value in overflow):
                return False, []

        normalized_first_header = self._normalize_headers(
            tuple(list(first_row)[:expected_width])
        )
        if normalized_first_header == expected_headers:
            return False, []

        normalized_rows = [
            self._normalize_row(
                row,
                expected_width,
                sheet_name=previous_sheet.sheet_name,
            )
            for row in populated_rows
        ]
        candidate_schema = self._infer_arrow_schema(
            expected_headers,
            normalized_rows,
        )
        previous_families = [
            self._dtype_family(previous_sheet.dtypes.get(column, ""))
            for column in expected_headers
        ]
        candidate_families = [
            self._dtype_family(str(field.type))
            for field in candidate_schema
        ]
        if candidate_families != previous_families:
            return False, []

        decisive_families = {"number", "datetime", "boolean"}
        compatible_first_row_fields = 0
        first_values = normalized_rows[0]
        for index, family in enumerate(previous_families):
            if family not in decisive_families:
                continue
            value = first_values[index]
            if value in (None, ""):
                continue
            first_value_type = self._infer_arrow_type([value])
            if self._dtype_family(str(first_value_type)) != family:
                return False, []
            compatible_first_row_fields += 1
        if compatible_first_row_fields == 0:
            return False, []

        return True, [
            "column count exactly matches the preceding worksheet",
            "sampled column type families exactly match the preceding worksheet",
            "the first physical row is data-compatible and is preserved as data",
        ]

    @staticmethod
    def _sheet_group_confidence(sheets: list[SheetMeta]) -> float:
        del sheets
        return 0.92

    def _process_small_workbook(
        self,
        path: Path,
        dataset_id: str,
        *,
        progress_callback: Callable[[dict[str, Any]], None] | None,
        cancel_callback: Callable[[], bool] | None,
    ) -> list[SheetMeta]:
        with pd.ExcelFile(path) as workbook:
            sheets: list[SheetMeta] = []
            sheet_names = list(workbook.sheet_names)
            for index, sheet_name in enumerate(sheet_names, start=1):
                self._raise_if_cancelled(cancel_callback)
                self._emit_progress(
                    progress_callback,
                    "profiling",
                    10 + int(index / max(1, len(sheet_names)) * 80),
                    sheet_name,
                )
                sheets.append(
                    self._process_sheet(
                        workbook,
                        sheet_name,
                        dataset_id,
                        previous_sheet=(sheets[-1] if sheets else None),
                    )
                )
            return sheets

    def _process_sheet(
        self,
        workbook: pd.ExcelFile,
        sheet_name: str,
        dataset_id: str,
        *,
        previous_sheet: SheetMeta | None = None,
    ) -> SheetMeta:
        inherited_header = False
        continuation_reasons: list[str] = []
        if previous_sheet is not None:
            preview = workbook.parse(
                sheet_name,
                header=None,
                nrows=max(4, self._schema_sample_rows + 1),
            )
            preview_rows = list(preview.itertuples(index=False, name=None))
            if preview_rows:
                inherited_header, continuation_reasons = (
                    self._detect_headerless_continuation(
                        previous_sheet,
                        preview_rows[0],
                        preview_rows[1:],
                    )
                )

        if inherited_header and previous_sheet is not None:
            dataframe = workbook.parse(
                sheet_name,
                header=None,
                names=list(previous_sheet.columns),
            )
            header_mode = "inherited"
            header_source_sheet_id = (
                previous_sheet.header_source_sheet_id
                or previous_sheet.sheet_id
                or previous_sheet.sheet_name
            )
            logger.info(
                "Inherited worksheet headers sheet=%s source=%s reasons=%s",
                sheet_name,
                previous_sheet.sheet_name,
                continuation_reasons,
            )
        else:
            dataframe = workbook.parse(sheet_name)
            header_mode = "own"
            header_source_sheet_id = ""
        dataframe, type_profiles = self._normalize_dataframe_for_parquet(
            dataframe,
            sheet_name=sheet_name,
        )
        sheet_hash = hashlib.sha256(f"{dataset_id}:{sheet_name}".encode("utf-8")).hexdigest()[:12]
        sheet_id = f"sh_{sheet_hash}"

        cache_dir = settings.DATASET_CACHE_DIR / dataset_id
        cache_dir.mkdir(parents=True, exist_ok=True)
        cache_path = cache_dir / f"{sheet_id}.parquet"
        sample_cache_path = cache_dir / f"{sheet_id}.sample.parquet"
        self._write_dataframe_parquet_atomically(dataframe, cache_path)
        self._write_dataframe_parquet_atomically(
            self._representative_sample(dataframe),
            sample_cache_path,
        )

        numeric_columns = dataframe.select_dtypes(include="number").columns[: self._max_cols_describe]
        describe_raw = dataframe[numeric_columns].describe().to_dict() if len(numeric_columns) else {}
        describe = {
            str(column): {
                str(name): round(float(value), 4)
                for name, value in statistics.items()
            }
            for column, statistics in describe_raw.items()
        }

        unique_counts = {
            str(column): int(dataframe[column].nunique(dropna=True))
            for column in dataframe.columns
        }
        denominator = max(1, len(dataframe))
        unique_rates = {
            column: round(count / denominator, 6)
            for column, count in unique_counts.items()
        }

        return SheetMeta(
            sheet_name=sheet_name,
            sheet_id=sheet_id,
            cache_path=str(cache_path),
            sample_cache_path=str(sample_cache_path),
            rows=len(dataframe),
            cols=len(dataframe.columns),
            columns=[str(column) for column in dataframe.columns],
            dtypes={str(column): str(dtype) for column, dtype in dataframe.dtypes.items()},
            null_counts={
                str(column): int(count)
                for column, count in dataframe.isnull().sum().items()
                if count > 0
            },
            head_sample=(
                dataframe.head(self._preview_rows).fillna("").astype(str).to_dict(orient="records")
            ),
            describe=describe,
            unique_values=self._collect_unique_values(dataframe),
            unique_counts=unique_counts,
            unique_rates=unique_rates,
            type_profiles=type_profiles,
            semantic_roles=self._detect_semantic_roles(
                [str(column) for column in dataframe.columns],
                {str(column): str(dtype) for column, dtype in dataframe.dtypes.items()},
                type_profiles,
            ),
            header_mode=header_mode,
            header_source_sheet_id=header_source_sheet_id,
            first_row_is_data=inherited_header,
            continuation_detected=inherited_header,
            continuation_confidence=(1.0 if inherited_header else 0.0),
            continuation_reasons=continuation_reasons,
        )

    def _process_large_xlsx(
        self,
        path: Path,
        dataset_id: str,
        *,
        progress_callback: Callable[[dict[str, Any]], None] | None = None,
        cancel_callback: Callable[[], bool] | None = None,
        import_plan: XlsxImportPlan | None = None,
    ) -> list[SheetMeta]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheets = []
            worksheets = list(workbook.worksheets)
            for index, worksheet in enumerate(worksheets, start=1):
                self._raise_if_cancelled(cancel_callback)
                self._emit_progress(
                    progress_callback,
                    "importing",
                    10 + int(index / max(1, len(worksheets)) * 65),
                    str(worksheet.title),
                )
                worksheet_path = str(
                    getattr(worksheet, "_worksheet_path", "")
                ).replace("\\", "/").lstrip("/")
                worksheet_info = (
                    import_plan.worksheet(worksheet_path)
                    if import_plan is not None
                    else None
                )
                try:
                    declared_dimension = worksheet.calculate_dimension()
                except ValueError:
                    declared_dimension = None

                # A worksheet dimension is optional metadata.  In read-only mode
                # openpyxl treats it as a hard boundary, so remove that boundary
                # and let the streaming parser consume the real sheetData rows.
                worksheet.reset_dimensions()
                try:
                    sheet_meta = self._stream_sheet_to_cache(
                        worksheet,
                        dataset_id,
                        cancel_callback=cancel_callback,
                        safe_mode=bool(import_plan and import_plan.mode == "safe"),
                        previous_sheet=(sheets[-1] if sheets else None),
                    )
                except Exception:
                    self._cleanup_dataset_partial_files(dataset_id)
                    raise
                self._validate_streamed_sheet(
                    sheet_meta,
                    declared_dimension=declared_dimension,
                    worksheet_info=worksheet_info,
                )
                sheets.append(sheet_meta)
                self._emit_progress(
                    progress_callback,
                    "profiling",
                    75 + int(index / max(1, len(worksheets)) * 20),
                    str(worksheet.title),
                )
            return sheets
        finally:
            workbook.close()

    def _stream_sheet_to_cache(
        self,
        worksheet: Any,
        dataset_id: str,
        *,
        cancel_callback: Callable[[], bool] | None,
        forced_text_columns: set[str] | None = None,
        safe_mode: bool = False,
        previous_sheet: SheetMeta | None = None,
    ) -> SheetMeta:
        forced_text_columns = set(forced_text_columns or ())
        sheet_name = str(worksheet.title)
        sheet_hash = hashlib.sha256(f"{dataset_id}:{sheet_name}".encode("utf-8")).hexdigest()[:12]
        sheet_id = f"sh_{sheet_hash}"
        cache_dir = settings.DATASET_CACHE_DIR / dataset_id
        cache_dir.mkdir(parents=True, exist_ok=True)

        cache_path = cache_dir / f"{sheet_id}.parquet"
        sample_cache_path = cache_dir / f"{sheet_id}.sample.parquet"
        temp_cache_path = cache_dir / f"{sheet_id}.partial.parquet"
        temp_sample_path = cache_dir / f"{sheet_id}.sample.partial.parquet"
        raw_temp_path = cache_dir / f"{sheet_id}.raw.partial.parquet"

        self._cleanup_partial_files(temp_cache_path, temp_sample_path, raw_temp_path)
        logger.info(
            "Streaming sheet %s into parquet cache %s mode=%s",
            sheet_name,
            cache_path,
            "safe" if safe_mode else "typed",
        )

        row_iterator = worksheet.iter_rows(values_only=True)
        first_row = next(row_iterator, None)
        if first_row is None:
            empty = pd.DataFrame()
            self._write_dataframe_parquet_atomically(empty, cache_path)
            self._write_dataframe_parquet_atomically(empty, sample_cache_path)
            return SheetMeta(
                sheet_name=sheet_name,
                sheet_id=sheet_id,
                cache_path=str(cache_path),
                sample_cache_path=str(sample_cache_path),
                rows=0,
                cols=0,
                columns=[],
                dtypes={},
                null_counts={},
                head_sample=[],
                describe={},
                unique_values={},
            )

        raw_seed_rows: list[Any] = []
        for _ in range(self._schema_sample_rows):
            raw = next(row_iterator, None)
            if raw is None:
                break
            raw_seed_rows.append(raw)

        inherited_header, continuation_reasons = (
            self._detect_headerless_continuation(
                previous_sheet,
                first_row,
                raw_seed_rows,
            )
        )
        if inherited_header and previous_sheet is not None:
            headers = list(previous_sheet.columns)
            width = len(headers)
            seed_rows = [
                self._normalize_row(raw, width, sheet_name=sheet_name)
                for raw in [first_row, *raw_seed_rows]
            ]
            header_mode = "inherited"
            header_source_sheet_id = (
                previous_sheet.header_source_sheet_id
                or previous_sheet.sheet_id
                or previous_sheet.sheet_name
            )
            first_row_is_data = True
            logger.info(
                "Inherited worksheet headers sheet=%s source=%s reasons=%s",
                sheet_name,
                previous_sheet.sheet_name,
                continuation_reasons,
            )
        else:
            width = max(
                len(first_row),
                *(len(row) for row in raw_seed_rows),
            )
            header_values = list(first_row)
            if len(header_values) < width:
                header_values.extend([None] * (width - len(header_values)))
            headers = self._normalize_headers(tuple(header_values))
            seed_rows = [
                self._normalize_row(raw, width, sheet_name=sheet_name)
                for raw in raw_seed_rows
            ]
            header_mode = "own"
            header_source_sheet_id = ""
            first_row_is_data = False

        if safe_mode:
            return self._stream_sheet_to_cache_safe(
                row_iterator=row_iterator,
                seed_rows=seed_rows,
                headers=headers,
                sheet_name=sheet_name,
                sheet_id=sheet_id,
                cache_path=cache_path,
                sample_cache_path=sample_cache_path,
                temp_cache_path=temp_cache_path,
                temp_sample_path=temp_sample_path,
                raw_temp_path=raw_temp_path,
                cancel_callback=cancel_callback,
                header_mode=header_mode,
                header_source_sheet_id=header_source_sheet_id,
                first_row_is_data=first_row_is_data,
                continuation_reasons=continuation_reasons,
            )

        schema = self._infer_arrow_schema(
            headers,
            seed_rows,
            forced_text_columns=forced_text_columns,
        )
        writer = pq.ParquetWriter(
            temp_cache_path,
            schema=schema,
            compression="zstd",
            use_dictionary=True,
        )
        rows_seen = 0
        null_counts = Counter()
        head_rows: list[dict[str, Any]] = []
        reservoir: list[dict[str, Any]] = []
        reservoir_rng = random.Random(42)
        text_uniques: dict[str, set[str]] = {header: set() for header in headers}

        contamination: SchemaContamination | None = None
        try:
            for batch in self._iter_normalized_batches(
                row_iterator,
                seed_rows,
                width=len(headers),
                sheet_name=sheet_name,
                cancel_callback=cancel_callback,
            ):
                table, records = self._batch_to_arrow_table(headers, schema, batch)
                writer.write_table(table, row_group_size=self._row_group_size)
                rows_seen = self._update_stream_profile(
                    records,
                    rows_seen,
                    null_counts,
                    head_rows,
                    reservoir,
                    reservoir_rng,
                    text_uniques,
                )
        except SchemaContamination as exc:
            contamination = exc
            logger.warning(
                "Type contamination detected dataset_id=%s sheet=%s column=%s value=%r; "
                "falling back once to raw safe import",
                dataset_id,
                sheet_name,
                exc.column,
                exc.value,
            )
        except Exception:
            logger.exception("Large-sheet import failed for %s/%s", dataset_id, sheet_name)
            raise
        finally:
            writer.close()

        if contamination is not None:
            self._cleanup_partial_files(temp_cache_path, temp_sample_path, raw_temp_path)
            worksheet.reset_dimensions()
            return self._stream_sheet_to_cache(
                worksheet,
                dataset_id,
                cancel_callback=cancel_callback,
                safe_mode=True,
                previous_sheet=previous_sheet,
            )

        sample_df = pd.DataFrame(reservoir, columns=headers) if reservoir else pd.DataFrame(columns=headers)
        return self._complete_streamed_sheet(
            sheet_name=sheet_name,
            sheet_id=sheet_id,
            cache_path=cache_path,
            sample_cache_path=sample_cache_path,
            temp_cache_path=temp_cache_path,
            temp_sample_path=temp_sample_path,
            schema=schema,
            sample_df=sample_df,
            rows=rows_seen,
            headers=headers,
            null_counts=null_counts,
            head_rows=head_rows,
            header_mode=header_mode,
            header_source_sheet_id=header_source_sheet_id,
            first_row_is_data=first_row_is_data,
            continuation_reasons=continuation_reasons,
        )

    def _stream_sheet_to_cache_safe(
        self,
        *,
        row_iterator: Any,
        seed_rows: list[list[Any]],
        headers: list[str],
        sheet_name: str,
        sheet_id: str,
        cache_path: Path,
        sample_cache_path: Path,
        temp_cache_path: Path,
        temp_sample_path: Path,
        raw_temp_path: Path,
        cancel_callback: Callable[[], bool] | None,
        header_mode: str,
        header_source_sheet_id: str,
        first_row_is_data: bool,
        continuation_reasons: list[str],
    ) -> SheetMeta:
        """Read Excel once into a raw cache, then type it from local Parquet."""
        raw_schema = pa.schema([pa.field(header, pa.string()) for header in headers])
        writer = pq.ParquetWriter(
            raw_temp_path,
            schema=raw_schema,
            compression="zstd",
            use_dictionary=True,
        )
        observed_kinds: list[set[str]] = [set() for _ in headers]
        rows_seen = 0
        null_counts = Counter()
        head_rows: list[dict[str, Any]] = []
        reservoir: list[dict[str, Any]] = []
        reservoir_rng = random.Random(42)
        text_uniques: dict[str, set[str]] = {header: set() for header in headers}
        try:
            for batch in self._iter_normalized_batches(
                row_iterator,
                seed_rows,
                width=len(headers),
                sheet_name=sheet_name,
                cancel_callback=cancel_callback,
            ):
                table, records = self._batch_to_raw_arrow_table(
                    headers,
                    batch,
                    observed_kinds,
                )
                writer.write_table(table, row_group_size=self._row_group_size)
                rows_seen = self._update_stream_profile(
                    records,
                    rows_seen,
                    null_counts,
                    head_rows,
                    reservoir,
                    reservoir_rng,
                    text_uniques,
                )
        except Exception:
            logger.exception("Safe streaming import failed for sheet=%s", sheet_name)
            raise
        finally:
            writer.close()

        schema = self._schema_from_observed_kinds(headers, observed_kinds)
        try:
            self._materialize_typed_parquet(
                raw_temp_path,
                temp_cache_path,
                schema,
            )
        finally:
            raw_temp_path.unlink(missing_ok=True)

        sample_df = (
            pd.DataFrame(reservoir, columns=headers)
            if reservoir
            else pd.DataFrame(columns=headers)
        )
        sample_df = self._coerce_sample_to_schema(sample_df, schema)
        return self._complete_streamed_sheet(
            sheet_name=sheet_name,
            sheet_id=sheet_id,
            cache_path=cache_path,
            sample_cache_path=sample_cache_path,
            temp_cache_path=temp_cache_path,
            temp_sample_path=temp_sample_path,
            schema=schema,
            sample_df=sample_df,
            rows=rows_seen,
            headers=headers,
            null_counts=null_counts,
            head_rows=head_rows,
            header_mode=header_mode,
            header_source_sheet_id=header_source_sheet_id,
            first_row_is_data=first_row_is_data,
            continuation_reasons=continuation_reasons,
        )

    def _complete_streamed_sheet(
        self,
        *,
        sheet_name: str,
        sheet_id: str,
        cache_path: Path,
        sample_cache_path: Path,
        temp_cache_path: Path,
        temp_sample_path: Path,
        schema: pa.Schema,
        sample_df: pd.DataFrame,
        rows: int,
        headers: list[str],
        null_counts: Counter,
        head_rows: list[dict[str, Any]],
        header_mode: str = "own",
        header_source_sheet_id: str = "",
        first_row_is_data: bool = False,
        continuation_reasons: list[str] | None = None,
    ) -> SheetMeta:
        self._write_dataframe_parquet_atomically(sample_df, temp_sample_path)
        unique_counts = self._approx_unique_counts(temp_cache_path, headers)
        unique_rates = {
            column: round(count / max(1, rows), 6)
            for column, count in unique_counts.items()
        }
        dtypes = {field.name: str(field.type) for field in schema}
        type_profiles = self._build_type_profiles(sample_df, dtypes)
        describe = self._describe_sample(sample_df, dtypes)
        unique_values = self._collect_unique_values(sample_df)

        self._finalize_atomic_cache(temp_cache_path, cache_path)
        self._finalize_atomic_cache(temp_sample_path, sample_cache_path)
        logger.info(
            "Completed streaming sheet %s rows=%s cols=%s",
            sheet_name,
            rows,
            len(headers),
        )
        return SheetMeta(
            sheet_name=sheet_name,
            sheet_id=sheet_id,
            cache_path=str(cache_path),
            sample_cache_path=str(sample_cache_path),
            rows=rows,
            cols=len(headers),
            columns=headers,
            dtypes=dtypes,
            null_counts={
                key: int(value)
                for key, value in null_counts.items()
                if value > 0
            },
            head_sample=[
                {
                    column: "" if value is None else str(value)
                    for column, value in row.items()
                }
                for row in head_rows
            ],
            describe=describe,
            unique_values=unique_values,
            unique_counts=unique_counts,
            unique_rates=unique_rates,
            type_profiles=type_profiles,
            semantic_roles=self._detect_semantic_roles(
                headers,
                dtypes,
                type_profiles,
            ),
            header_mode=header_mode,
            header_source_sheet_id=header_source_sheet_id,
            first_row_is_data=first_row_is_data,
            continuation_detected=(header_mode == "inherited"),
            continuation_confidence=(1.0 if header_mode == "inherited" else 0.0),
            continuation_reasons=list(continuation_reasons or []),
        )

    def _iter_normalized_batches(
        self,
        row_iterator: Any,
        seed_rows: list[list[Any]],
        *,
        width: int,
        sheet_name: str,
        cancel_callback: Callable[[], bool] | None,
    ) -> Any:
        """Yield memory-budgeted row batches without buffering the worksheet."""
        pending = deque(seed_rows)
        carry: list[Any] | None = None
        source_exhausted = False
        target_bytes = self._effective_batch_target_bytes()
        max_rows = max(1, int(self._batch_rows))
        min_rows = max(1, min(int(self._min_batch_rows), max_rows))
        cancel_interval = max(1, int(self._cancel_check_rows))
        rows_since_cancel = 0

        while pending or carry is not None or not source_exhausted:
            batch: list[list[Any]] = []
            estimated_bytes = 0
            while len(batch) < max_rows:
                if carry is not None:
                    normalized = carry
                    carry = None
                elif pending:
                    normalized = pending.popleft()
                elif not source_exhausted:
                    raw = next(row_iterator, None)
                    if raw is None:
                        source_exhausted = True
                        break
                    normalized = self._normalize_row(
                        raw,
                        width,
                        sheet_name=sheet_name,
                    )
                else:
                    break

                row_bytes = self._estimate_row_memory_bytes(normalized)
                would_exceed = estimated_bytes + row_bytes > target_bytes
                large_row = row_bytes > max(1, target_bytes // 4)
                if batch and would_exceed and (len(batch) >= min_rows or large_row):
                    carry = normalized
                    break

                batch.append(normalized)
                estimated_bytes += row_bytes
                rows_since_cancel += 1
                if rows_since_cancel >= cancel_interval:
                    self._raise_if_cancelled(cancel_callback)
                    rows_since_cancel = 0

            if not batch:
                break
            self._raise_if_cancelled(cancel_callback)
            yield batch

    def _effective_batch_target_bytes(self) -> int:
        available_memory = max(1, int(psutil.virtual_memory().available))
        fraction = max(0.01, min(0.50, float(self._batch_memory_fraction)))
        memory_budget = max(64 * 1024, int(available_memory * fraction))
        configured = max(64 * 1024, int(self._batch_target_bytes))
        return min(configured, memory_budget)

    @staticmethod
    def _estimate_row_memory_bytes(row: list[Any]) -> int:
        estimated = 64
        for value in row:
            if value is None:
                estimated += 8
            elif isinstance(value, str):
                estimated += 64 + len(value.encode("utf-8", errors="replace"))
            elif isinstance(value, (datetime, date, time)):
                estimated += 64
            else:
                estimated += 32
        # Lists, records, column buffers, and Arrow arrays coexist briefly.
        return max(1, int(estimated * 2.5))

    def _batch_to_raw_arrow_table(
        self,
        headers: list[str],
        rows: list[list[Any]],
        observed_kinds: list[set[str]],
    ) -> tuple[pa.Table, list[dict[str, Any]]]:
        columns: dict[str, list[str | None]] = {header: [] for header in headers}
        records: list[dict[str, Any]] = []
        for row in rows:
            record: dict[str, Any] = {}
            for index, header in enumerate(headers):
                value = row[index]
                kind = self._value_kind(value)
                if kind is not None:
                    observed_kinds[index].add(kind)
                raw_value = self._raw_scalar(value)
                columns[header].append(raw_value)
                record[header] = raw_value
            records.append(record)
        arrays = [pa.array(columns[header], type=pa.string()) for header in headers]
        return pa.Table.from_arrays(arrays, names=headers), records

    @staticmethod
    def _value_kind(value: Any) -> str | None:
        if value is None or (isinstance(value, str) and not value.strip()):
            return None
        if isinstance(value, bool):
            return "bool"
        if isinstance(value, int):
            return "int"
        if isinstance(value, float):
            return "float"
        if isinstance(value, datetime):
            return "datetime"
        if isinstance(value, date):
            return "date"
        if isinstance(value, time):
            return "time"
        return "string"

    @staticmethod
    def _raw_scalar(value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            stripped = value.strip()
            return stripped or None
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, (datetime, date, time)):
            return value.isoformat()
        return str(value)

    @staticmethod
    def _schema_from_observed_kinds(
        headers: list[str],
        observed_kinds: list[set[str]],
    ) -> pa.Schema:
        fields: list[pa.Field] = []
        for header, kinds in zip(headers, observed_kinds):
            if kinds and kinds <= {"bool"}:
                arrow_type = pa.bool_()
            elif kinds and kinds <= {"int"}:
                arrow_type = pa.int64()
            elif kinds and kinds <= {"int", "float"}:
                arrow_type = pa.float64()
            elif kinds and kinds <= {"datetime"}:
                arrow_type = pa.timestamp("us")
            elif kinds and kinds <= {"date"}:
                arrow_type = pa.date32()
            elif kinds and kinds <= {"date", "datetime"}:
                arrow_type = pa.timestamp("us")
            else:
                arrow_type = pa.string()
            fields.append(pa.field(header, arrow_type))
        return pa.schema(fields)

    def _materialize_typed_parquet(
        self,
        raw_path: Path,
        destination: Path,
        schema: pa.Schema,
    ) -> None:
        destination.unlink(missing_ok=True)
        if all(pa.types.is_string(field.type) for field in schema):
            raw_path.replace(destination)
            return

        sql_types: dict[str, str] = {}
        for field in schema:
            if pa.types.is_boolean(field.type):
                sql_types[field.name] = "BOOLEAN"
            elif pa.types.is_integer(field.type):
                sql_types[field.name] = "BIGINT"
            elif pa.types.is_floating(field.type):
                sql_types[field.name] = "DOUBLE"
            elif pa.types.is_timestamp(field.type):
                sql_types[field.name] = "TIMESTAMP"
            elif pa.types.is_date32(field.type):
                sql_types[field.name] = "DATE"

        projections = []
        for field in schema:
            quoted = self._quote(field.name)
            sql_type = sql_types.get(field.name)
            projections.append(
                f"CAST({quoted} AS {sql_type}) AS {quoted}"
                if sql_type
                else quoted
            )
        raw_sql = str(raw_path).replace("'", "''")
        destination_sql = str(destination).replace("'", "''")
        connection = self._duckdb_connection()
        try:
            connection.execute("SET preserve_insertion_order=true")
            connection.execute(
                "COPY (SELECT "
                + ",".join(projections)
                + f" FROM read_parquet('{raw_sql}')) "
                + f"TO '{destination_sql}' "
                + "(FORMAT PARQUET, COMPRESSION ZSTD, "
                + f"ROW_GROUP_SIZE {max(1, int(self._row_group_size))})"
            )
        finally:
            connection.close()

    @staticmethod
    def _coerce_sample_to_schema(
        sample: pd.DataFrame,
        schema: pa.Schema,
    ) -> pd.DataFrame:
        converted = sample.copy()
        for field in schema:
            column = field.name
            if column not in converted:
                continue
            if pa.types.is_boolean(field.type):
                converted[column] = converted[column].map(
                    lambda value: (
                        pd.NA
                        if pd.isna(value)
                        else str(value).strip().lower() == "true"
                    )
                ).astype("boolean")
            elif pa.types.is_integer(field.type):
                converted[column] = pd.to_numeric(
                    converted[column],
                    errors="raise",
                ).astype("Int64")
            elif pa.types.is_floating(field.type):
                converted[column] = pd.to_numeric(
                    converted[column],
                    errors="raise",
                ).astype("Float64")
            elif pa.types.is_timestamp(field.type):
                converted[column] = pd.to_datetime(converted[column], errors="raise")
            elif pa.types.is_date32(field.type):
                converted[column] = pd.to_datetime(
                    converted[column],
                    errors="raise",
                ).dt.date
            else:
                converted[column] = converted[column].astype("string")
        return converted

    @staticmethod
    def _validate_streamed_sheet(
        sheet: SheetMeta,
        *,
        declared_dimension: str | None,
        worksheet_info: XlsxWorksheetInfo | None,
    ) -> None:
        if (
            worksheet_info is not None
            and worksheet_info.has_data_beyond_a1
            and sheet.rows == 0
            and sheet.cols <= 1
        ):
            raise ValueError(
                f"Worksheet {sheet.sheet_name!r} contains cells beyond A1 but "
                "the streaming import produced no data rows."
            )

        if sheet.cols <= 0:
            actual_dimension = "<empty>"
        else:
            physical_rows = sheet.rows + (0 if sheet.first_row_is_data else 1)
            actual_dimension = (
                f"A1:{get_column_letter(sheet.cols)}{max(1, physical_rows)}"
            )
        normalized_declared = (declared_dimension or "<missing>").replace("$", "")
        if normalized_declared not in {actual_dimension, "A1" if actual_dimension == "A1:A1" else ""}:
            logger.warning(
                "Worksheet dimension mismatch sheet=%s declared=%s actual=%s; "
                "actual streamed cells were retained",
                sheet.sheet_name,
                normalized_declared,
                actual_dimension,
            )

    @staticmethod
    def _should_stream_xlsx(
        path: Path,
        size_kb: float,
        *,
        cancel_callback: Callable[[], bool] | None = None,
        import_plan: XlsxImportPlan | None = None,
    ) -> bool:
        if import_plan is not None:
            return import_plan.mode in {"stream", "safe"}
        if size_kb >= settings.LARGE_EXCEL_MB * 1024:
            return True
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                if cancel_callback and cancel_callback():
                    raise ProcessingCancelled("Dataset import cancelled")
                if int(worksheet.max_row or 0) >= settings.LARGE_DATASET_ROWS:
                    return True
            return False
        finally:
            workbook.close()

    def _ensure_free_disk_space(
        self,
        path: Path,
        *,
        import_plan: XlsxImportPlan | None = None,
    ) -> None:
        try:
            usage = shutil.disk_usage(settings.DATASET_CACHE_DIR)
        except FileNotFoundError:
            settings.DATASET_CACHE_DIR.mkdir(parents=True, exist_ok=True)
            usage = shutil.disk_usage(settings.DATASET_CACHE_DIR)

        estimated_need = self._estimate_required_disk_bytes(
            path,
            uncompressed_bytes=(
                import_plan.uncompressed_bytes
                if import_plan is not None
                else None
            ),
        )
        logger.info(
            "Disk check for %s free=%s required=%s",
            path.name,
            usage.free,
            estimated_need,
        )
        if usage.free < estimated_need:
            required_gb = estimated_need / (1024**3)
            free_gb = usage.free / (1024**3)
            raise ValueError(
                "Not enough free disk space for import. "
                f"Need about {required_gb:.1f} GB free, found {free_gb:.1f} GB."
            )

    @staticmethod
    def _estimate_required_disk_bytes(
        path: Path,
        *,
        uncompressed_bytes: int | None = None,
    ) -> int:
        source_size = path.stat().st_size
        zip_uncompressed = int(uncompressed_bytes or 0)
        if zip_uncompressed <= 0 and path.suffix.lower() in {".xlsx", ".xlsm"}:
            try:
                with zipfile.ZipFile(path) as archive:
                    zip_uncompressed = sum(info.file_size for info in archive.infolist())
            except zipfile.BadZipFile:
                logger.warning("Could not inspect ZIP members for %s", path)
        estimated = max(
            settings.IMPORT_MIN_FREE_DISK_BYTES,
            int(source_size * settings.IMPORT_SOURCE_SIZE_MULTIPLIER),
            int(zip_uncompressed * settings.IMPORT_UNCOMPRESSED_MULTIPLIER),
        )
        return estimated

    @staticmethod
    def _normalize_headers(values: tuple[Any, ...]) -> list[str]:
        seen: dict[str, int] = {}
        headers: list[str] = []
        for index, value in enumerate(values, start=1):
            base = str(value).strip() if value is not None else ""
            base = base or f"Column_{index}"
            seen[base] = seen.get(base, 0) + 1
            headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
        return headers

    @staticmethod
    def _normalize_row(
        values: Any,
        width: int,
        *,
        sheet_name: str = "",
    ) -> list[Any]:
        raw = list(values) if values is not None else []
        overflow = raw[width:]
        if any(value not in (None, "") for value in overflow):
            location = f" in worksheet {sheet_name!r}" if sheet_name else ""
            raise ValueError(
                "A data row contains populated cells beyond the detected header "
                f"width{location}; refusing to silently truncate columns."
            )
        row = raw[:width]
        if len(row) < width:
            row.extend([None] * (width - len(row)))
        return row

    def _batch_to_arrow_table(
        self,
        headers: list[str],
        schema: pa.Schema,
        rows: list[list[Any]],
    ) -> tuple[pa.Table, list[dict[str, Any]]]:
        records: list[dict[str, Any]] = []
        columns: dict[str, list[Any]] = {header: [] for header in headers}
        for row in rows:
            record: dict[str, Any] = {}
            for index, header in enumerate(headers):
                field = schema.field(index)
                value = self._coerce_scalar(
                    row[index],
                    field.type,
                    column=header,
                )
                columns[header].append(value)
                record[header] = value
            records.append(record)
        arrays = [pa.array(columns[field.name], type=field.type) for field in schema]
        return pa.Table.from_arrays(arrays, schema=schema), records

    def _infer_arrow_schema(
        self,
        headers: list[str],
        sample_rows: list[list[Any]],
        *,
        forced_text_columns: set[str] | None = None,
    ) -> pa.Schema:
        forced_text_columns = forced_text_columns or set()
        fields = []
        for index, header in enumerate(headers):
            observed = [row[index] for row in sample_rows if row[index] not in (None, "")]
            arrow_type = (
                pa.string()
                if header in forced_text_columns
                else self._infer_arrow_type(observed)
            )
            fields.append(pa.field(header, arrow_type))
        return pa.schema(fields)

    @staticmethod
    def _infer_arrow_type(values: list[Any]) -> pa.DataType:
        if not values:
            return pa.string()
        kinds = set()
        for value in values:
            if isinstance(value, bool):
                kinds.add("bool")
            elif isinstance(value, int) and not isinstance(value, bool):
                kinds.add("int")
            elif isinstance(value, float):
                kinds.add("float")
            elif isinstance(value, datetime):
                kinds.add("datetime")
            elif isinstance(value, date):
                kinds.add("date")
            elif isinstance(value, time):
                kinds.add("time")
            else:
                kinds.add("string")
        if kinds <= {"bool"}:
            return pa.bool_()
        if kinds <= {"bool", "int"}:
            return pa.int64()
        if kinds <= {"bool", "int", "float"}:
            return pa.float64()
        if kinds <= {"datetime"}:
            return pa.timestamp("us")
        if kinds <= {"date"}:
            return pa.date32()
        return pa.string()

    @staticmethod
    def _coerce_scalar(
        value: Any,
        arrow_type: pa.DataType,
        *,
        column: str = "",
    ) -> Any:
        if value is None:
            return None
        if isinstance(value, str):
            stripped = value.strip()
            if stripped == "":
                return None
            value = stripped
        try:
            if pa.types.is_boolean(arrow_type):
                if isinstance(value, str):
                    normalized = value.lower()
                    if normalized in {"true", "1", "yes", "y"}:
                        return True
                    if normalized in {"false", "0", "no", "n"}:
                        return False
                    raise SchemaContamination(column, value)
                return bool(value)
            if pa.types.is_integer(arrow_type):
                if isinstance(value, bool):
                    return int(value)
                if isinstance(value, (int, float)) and math.isfinite(float(value)):
                    return int(value)
                raise SchemaContamination(column, value)
            if pa.types.is_floating(arrow_type):
                if isinstance(value, bool):
                    return float(int(value))
                if isinstance(value, (int, float)) and math.isfinite(float(value)):
                    return float(value)
                raise SchemaContamination(column, value)
            if pa.types.is_timestamp(arrow_type):
                if isinstance(value, datetime):
                    return value
                if isinstance(value, date):
                    return datetime.combine(value, time.min)
                raise SchemaContamination(column, value)
            if pa.types.is_date32(arrow_type):
                if isinstance(value, datetime):
                    return value.date()
                if isinstance(value, date):
                    return value
                raise SchemaContamination(column, value)
            if isinstance(value, (datetime, date, time)):
                return value.isoformat()
            return str(value)
        except SchemaContamination:
            raise
        except Exception:
            return None

    def _normalize_dataframe_for_parquet(
        self,
        dataframe: pd.DataFrame,
        *,
        sheet_name: str,
    ) -> tuple[pd.DataFrame, dict[str, dict[str, Any]]]:
        normalized = dataframe.copy()
        for column in normalized.select_dtypes(include=["object", "string"]).columns:
            series = normalized[column]
            normalized[column] = series.map(
                lambda value: None if pd.isna(value) else str(value)
            ).astype("string")
        profiles = self._build_type_profiles(
            normalized,
            {str(column): str(dtype) for column, dtype in normalized.dtypes.items()},
        )
        polluted = [
            column
            for column, profile in profiles.items()
            if int(profile.get("invalid_count", 0)) > 0
        ]
        if polluted:
            logger.warning(
                "Preserved mixed-type values as text sheet=%s columns=%s",
                sheet_name,
                polluted,
            )
        return normalized, profiles

    @staticmethod
    def _build_type_profiles(
        dataframe: pd.DataFrame,
        dtypes: dict[str, str],
    ) -> dict[str, dict[str, Any]]:
        profiles: dict[str, dict[str, Any]] = {}
        for column in dataframe.columns:
            dtype = str(dtypes.get(str(column), ""))
            lowered = dtype.lower()
            if not any(token in lowered for token in ("object", "string", "str")):
                continue
            series = dataframe[column]
            non_null = series.dropna()
            if non_null.empty:
                continue
            text = non_null.astype(str).str.strip()
            non_empty = text[text != ""]
            if non_empty.empty:
                continue
            numeric = pd.to_numeric(
                non_empty.str.replace(",", "", regex=False),
                errors="coerce",
            )
            valid_count = int(numeric.notna().sum())
            invalid = non_empty[numeric.isna()]
            invalid_count = int(len(invalid))
            rate = valid_count / max(1, len(non_empty))
            if valid_count < 2 or rate < 0.6:
                continue
            profiles[str(column)] = {
                "storage_type": "string",
                "inferred_type": "numeric",
                "non_empty_count": int(len(non_empty)),
                "valid_count": valid_count,
                "invalid_count": invalid_count,
                "conversion_rate": round(rate, 6),
                "invalid_examples": list(dict.fromkeys(invalid.astype(str).tolist()))[:5],
            }
        return profiles

    @classmethod
    def _detect_semantic_roles(
        cls,
        columns: list[str],
        dtypes: dict[str, str],
        type_profiles: dict[str, dict[str, Any]],
    ) -> dict[str, list[str]]:
        roles: dict[str, list[str]] = {}
        for column in columns:
            column_roles = cls._semantic_roles_for_column(
                column,
                dtypes.get(column, ""),
                type_profiles.get(column, {}),
            )
            for role in column_roles:
                roles.setdefault(role, []).append(column)
        return roles

    @classmethod
    def _semantic_roles_for_column(
        cls,
        column: str,
        dtype: str,
        type_profile: dict[str, Any],
    ) -> list[str]:
        raw = str(column).strip().lower()
        compact = "".join(character for character in raw if character.isalnum())
        dtype_lower = str(dtype).lower()
        numeric_like = cls._is_numeric_dtype(dtype_lower) or (
            type_profile.get("inferred_type") == "numeric"
        )
        datetime_like = any(
            token in dtype_lower
            for token in ("date", "time", "timestamp", "datetime")
        )
        roles: list[str] = []

        def add(role: str) -> None:
            if role not in roles:
                roles.append(role)

        if datetime_like or cls._matches_any(
            raw,
            compact,
            (
                "date",
                "postingdate",
                "documentdate",
                "transactiondate",
                "createddate",
                "日期",
                "入账日期",
                "过账日期",
                "凭证日期",
            ),
        ):
            add("date")
        if cls._matches_any(
            raw,
            compact,
            (
                "period",
                "fiscalperiod",
                "fiscalyear",
                "yearmonth",
                "month",
                "期间",
                "会计期间",
                "年月",
                "月份",
            ),
        ):
            add("period")
        if cls._matches_any(
            raw,
            compact,
            (
                "account",
                "accountcode",
                "accountnumber",
                "glaccount",
                "g/l account",
                "acct",
                "科目",
                "会计科目",
                "总账科目",
                "账户",
            ),
        ):
            add("account")
        if numeric_like and cls._matches_any(
            raw,
            compact,
            (
                "amount",
                "amt",
                "value",
                "balance",
                "金额",
                "余额",
                "本币金额",
                "原币金额",
            ),
        ):
            add("amount")
        if numeric_like and cls._matches_any(
            raw,
            compact,
            ("debit", "dramount", "debitamount", "借方", "借方金额"),
        ):
            add("debit")
            add("amount")
        if numeric_like and cls._matches_any(
            raw,
            compact,
            ("credit", "cramount", "creditamount", "贷方", "贷方金额"),
        ):
            add("credit")
            add("amount")
        if cls._matches_any(
            raw,
            compact,
            (
                "document",
                "documentno",
                "documentnumber",
                "docno",
                "voucherno",
                "journalentry",
                "entryid",
                "凭证",
                "凭证号",
                "单据",
                "分录",
            ),
        ):
            add("document")
        if cls._matches_any(
            raw,
            compact,
            ("vendor", "supplier", "供应商", "供货商"),
        ):
            add("vendor")
        if cls._matches_any(
            raw,
            compact,
            ("customer", "client", "客户"),
        ):
            add("customer")
        if cls._matches_any(
            raw,
            compact,
            ("company", "companycode", "entity", "bukrs", "公司", "公司代码", "法人"),
        ):
            add("company")
        if cls._matches_any(
            raw,
            compact,
            ("costcenter", "cost centre", "成本中心"),
        ):
            add("cost_center")
        if cls._matches_any(
            raw,
            compact,
            ("profitcenter", "利润中心"),
        ):
            add("profit_center")
        if cls._matches_any(
            raw,
            compact,
            ("currency", "curr", "币种", "货币"),
        ):
            add("currency")
        if cls._matches_any(
            raw,
            compact,
            (
                "description",
                "text",
                "memo",
                "摘要",
                "说明",
                "描述",
            ),
        ):
            add("description")
        if cls._matches_any(
            raw,
            compact,
            (
                "createdby",
                "preparedby",
                "postedby",
                "user",
                "operator",
                "制单人",
                "录入人",
                "创建人",
                "用户",
            ),
        ):
            add("user")
        return roles

    @staticmethod
    def _matches_any(raw: str, compact: str, patterns: tuple[str, ...]) -> bool:
        for pattern in patterns:
            normalized = pattern.lower()
            compact_pattern = "".join(
                character
                for character in normalized
                if character.isalnum()
            )
            if normalized and normalized in raw:
                return True
            if compact_pattern and compact_pattern in compact:
                return True
        return False

    @staticmethod
    def _is_numeric_dtype(dtype: str) -> bool:
        return any(
            token in dtype
            for token in ("int", "float", "double", "decimal", "number")
        )

    def _update_stream_profile(
        self,
        records: list[dict[str, Any]],
        rows_seen: int,
        null_counts: Counter,
        head_rows: list[dict[str, Any]],
        reservoir: list[dict[str, Any]],
        reservoir_rng: random.Random,
        text_uniques: dict[str, set[str]],
    ) -> int:
        for record in records:
            rows_seen += 1
            for column, value in record.items():
                if value is None:
                    null_counts[column] += 1
                elif isinstance(value, str) and len(text_uniques[column]) <= self._max_unique_values:
                    text_uniques[column].add(value)
            if len(head_rows) < self._preview_rows:
                head_rows.append(record.copy())
            if len(reservoir) < self._sample_rows:
                reservoir.append(record.copy())
            else:
                replacement_index = reservoir_rng.randint(0, rows_seen - 1)
                if replacement_index < self._sample_rows:
                    reservoir[replacement_index] = record.copy()
        return rows_seen

    def _approx_unique_counts(self, parquet_path: Path, columns: list[str]) -> dict[str, int]:
        if not columns:
            return {}
        connection = self._duckdb_connection()
        try:
            aggregates = [
                f"APPROX_COUNT_DISTINCT({self._quote(column)})"
                for column in columns
            ]
            values = connection.execute(
                "SELECT " + ",".join(aggregates) + " FROM read_parquet(?)",
                [str(parquet_path)],
            ).fetchone()
            return {
                column: int((values[index] or 0) if values else 0)
                for index, column in enumerate(columns)
            }
        finally:
            connection.close()

    def _describe_sample(self, sample: pd.DataFrame, dtypes: dict[str, str]) -> dict[str, Any]:
        if sample.empty:
            return {}
        numeric_columns = []
        for column, dtype in dtypes.items():
            upper = dtype.upper()
            if any(token in upper for token in ("INT", "DOUBLE", "FLOAT", "DECIMAL")):
                numeric_columns.append(column)
        numeric_columns = numeric_columns[: self._max_cols_describe]
        if not numeric_columns:
            return {}
        describe = sample[numeric_columns].describe().to_dict()
        return {
            str(column): {
                str(name): round(float(value), 4)
                for name, value in statistics.items()
                if pd.notna(value)
            }
            for column, statistics in describe.items()
        }

    def _collect_unique_values(self, dataframe: pd.DataFrame) -> dict[str, list[str]]:
        values: dict[str, list[str]] = {}
        text_columns = dataframe.select_dtypes(include=["object", "string"]).columns
        for column in text_columns:
            series = dataframe[column].dropna().astype(str)
            unique = sorted(series.unique().tolist())
            if 0 < len(unique) <= self._max_unique_values:
                values[str(column)] = unique
        return values

    @staticmethod
    def _representative_sample(dataframe: pd.DataFrame) -> pd.DataFrame:
        target = max(1, settings.SAMPLE_ROWS_PER_SHEET)
        if len(dataframe) <= target:
            return dataframe.copy()
        edge = min(1000, target // 4)
        edge_indices = set(dataframe.head(edge).index) | set(dataframe.tail(edge).index)
        remaining = dataframe.loc[~dataframe.index.isin(edge_indices)]
        random_count = min(max(0, target - len(edge_indices)), len(remaining))
        sampled = pd.concat(
            [
                dataframe.loc[sorted(edge_indices)],
                remaining.sample(n=random_count, random_state=42),
            ]
        )
        return sampled.head(target)

    @staticmethod
    def _write_dataframe_parquet_atomically(dataframe: pd.DataFrame, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_suffix(path.suffix + ".partial")
        temporary.unlink(missing_ok=True)
        dataframe.to_parquet(temporary, index=False)
        temporary.replace(path)

    @staticmethod
    def _finalize_atomic_cache(temporary: Path, final_path: Path) -> None:
        final_path.parent.mkdir(parents=True, exist_ok=True)
        final_path.unlink(missing_ok=True)
        temporary.replace(final_path)

    @staticmethod
    def _cleanup_partial_files(*paths: Path) -> None:
        for path in paths:
            path.unlink(missing_ok=True)

    @staticmethod
    def _cleanup_dataset_partial_files(dataset_id: str) -> None:
        cache_dir = settings.DATASET_CACHE_DIR / dataset_id
        if not cache_dir.exists():
            return
        for path in cache_dir.glob("*.partial*"):
            path.unlink(missing_ok=True)

    def _write_dataset_manifest(self, dataset_id: str, file_meta: FileMeta) -> None:
        manifest_path = settings.DATASET_CACHE_DIR / dataset_id / "manifest.json"
        manifest_path.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "dataset_id": file_meta.dataset_id,
            "file_name": file_meta.file_name,
            "file_path": file_meta.file_path,
            "source_fingerprint": file_meta.source_fingerprint,
            "content_hash": file_meta.content_hash,
            "schema_family_id": file_meta.schema_family_id,
            "profile_mode": file_meta.profile_mode,
            "profile_sample_rows": file_meta.profile_sample_rows,
            "sheet_count": file_meta.sheet_count,
            "sheet_groups": [
                group.to_prompt_dict()
                for group in file_meta.sheet_groups
            ],
            "sheets": [sheet.to_prompt_dict() for sheet in file_meta.sheets],
        }
        manifest_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    @staticmethod
    def _fingerprint(path: Path) -> str:
        stat = path.stat()
        payload = f"{path.resolve()}|{stat.st_size}|{stat.st_mtime_ns}".encode("utf-8")
        return hashlib.sha256(payload).hexdigest()

    @classmethod
    def _content_hash(
        cls,
        path: Path,
        *,
        cancel_callback: Callable[[], bool] | None = None,
    ) -> str:
        """Return a path-independent identity for the source workbook."""
        digest = hashlib.sha256()
        with path.open("rb") as source:
            while chunk := source.read(8 * 1024 * 1024):
                cls._raise_if_cancelled(cancel_callback)
                digest.update(chunk)
        return digest.hexdigest()

    @classmethod
    def schema_family_id(cls, sheets: list[SheetMeta]) -> str:
        """Identify reusable workbook structure without using names or values."""
        signatures = set()
        for sheet in sheets:
            columns = sorted(
                [
                    {
                        "name": " ".join(str(column).split()).casefold(),
                        "type": cls._dtype_family(sheet.dtypes.get(column, "")),
                        "roles": sorted(
                            role
                            for role, role_columns in sheet.semantic_roles.items()
                            if column in role_columns
                        ),
                    }
                    for column in sheet.columns
                ],
                key=lambda item: item["name"],
            )
            signatures.add(
                json.dumps(
                    columns,
                    ensure_ascii=False,
                    sort_keys=True,
                    separators=(",", ":"),
                )
            )
        canonical = json.dumps(
            sorted(signatures),
            ensure_ascii=False,
            separators=(",", ":"),
        )
        return "sf_" + hashlib.sha256(canonical.encode("utf-8")).hexdigest()[:24]

    @staticmethod
    def _quote(identifier: str) -> str:
        return '"' + str(identifier).replace('"', '""') + '"'

    @staticmethod
    def _emit_progress(
        callback: Callable[[dict[str, Any]], None] | None,
        stage: str,
        percent: int,
        detail: str,
    ) -> None:
        if callback is not None:
            callback(
                {
                    "stage": stage,
                    "percent": max(0, min(100, int(percent))),
                    "detail": detail,
                }
            )

    @staticmethod
    def _raise_if_cancelled(cancel_callback: Callable[[], bool] | None) -> None:
        if cancel_callback and cancel_callback():
            logger.info("Dataset import cancelled by user")
            raise ProcessingCancelled("Dataset import cancelled")

    @staticmethod
    def _duckdb_connection() -> duckdb.DuckDBPyConnection:
        settings.DUCKDB_TEMP_DIR.mkdir(parents=True, exist_ok=True)
        connection = duckdb.connect()
        temp_dir = str(settings.DUCKDB_TEMP_DIR).replace("'", "''")
        connection.execute(f"SET memory_limit='{settings.DUCKDB_MEMORY_LIMIT}'")
        connection.execute(f"SET threads={max(1, settings.DUCKDB_THREADS)}")
        connection.execute(f"SET temp_directory='{temp_dir}'")
        connection.execute(f"SET max_temp_directory_size='{settings.DUCKDB_MAX_TEMP_SIZE}'")
        connection.execute(
            "SET preserve_insertion_order="
            + ("true" if settings.DUCKDB_PRESERVE_INSERTION_ORDER == "true" else "false")
        )
        return connection
