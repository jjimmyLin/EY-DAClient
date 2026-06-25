"""Export structured analysis results to a readable Excel workbook."""

from __future__ import annotations

import base64
import io
import logging
import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorksheetImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.analysis_result import AnalysisResult


logger = logging.getLogger(__name__)

EXCEL_MAX_CELL_CHARACTERS = 32_767
EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLUMNS = 16_384
FORMULA_PREFIXES = ("=", "+", "-", "@")

HEADER_FILL = PatternFill("solid", fgColor="1A73E8")
SECTION_FILL = PatternFill("solid", fgColor="E8F0FE")
SUBTLE_FILL = PatternFill("solid", fgColor="F8F9FA")
WARNING_FILL = PatternFill("solid", fgColor="FFF4CE")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="202124", bold=True, size=18)
SECTION_FONT = Font(color="202124", bold=True, size=12)
LABEL_FONT = Font(color="5F6368", bold=True)
BODY_FONT = Font(color="202124")
THIN_GRAY = Side(style="thin", color="DADCE0")


class AnalysisExportService:
    """Create a bounded, presentation-ready workbook from an AnalysisResult."""

    def export_excel(
        self,
        result: AnalysisResult,
        output_path: str | Path,
        *,
        metadata: dict[str, Any] | None = None,
        cancel_callback: Callable[[], bool] | None = None,
    ) -> str:
        destination = Path(output_path)
        if destination.suffix.lower() != ".xlsx":
            destination = destination.with_suffix(".xlsx")
        destination.parent.mkdir(parents=True, exist_ok=True)
        partial = destination.with_suffix(destination.suffix + ".partial")
        partial.unlink(missing_ok=True)

        logger.info(
            "Analysis export started output=%s metrics=%s tables=%s charts=%s",
            destination,
            len(result.metrics),
            len(result.tables),
            len(result.charts),
        )
        try:
            self._check_cancel(cancel_callback)
            workbook = Workbook()
            summary_sheet = workbook.active
            summary_sheet.title = "Summary"
            self._write_summary(summary_sheet, result, metadata or {})

            used_names = {summary_sheet.title}
            for index, table in enumerate(result.tables, start=1):
                self._check_cancel(cancel_callback)
                title = self._unique_sheet_name(
                    table.title or f"Table {index}",
                    used_names,
                )
                used_names.add(title)
                sheet = workbook.create_sheet(title)
                self._write_table(sheet, table)

            if result.charts:
                self._check_cancel(cancel_callback)
                chart_sheet = workbook.create_sheet(
                    self._unique_sheet_name("Charts", used_names)
                )
                used_names.add(chart_sheet.title)
                self._write_charts(chart_sheet, result, cancel_callback)

            if result.audit:
                self._check_cancel(cancel_callback)
                audit_sheet = workbook.create_sheet(
                    self._unique_sheet_name("Audit", used_names)
                )
                self._write_audit(audit_sheet, result.audit)

            workbook.save(partial)
            partial.replace(destination)
        except Exception:
            partial.unlink(missing_ok=True)
            logger.exception("Analysis export failed output=%s", destination)
            raise

        logger.info("Analysis export completed output=%s", destination)
        return str(destination)

    def _write_summary(
        self,
        sheet,
        result: AnalysisResult,
        metadata: dict[str, Any],
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A3"
        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 28
        sheet.column_dimensions["C"].width = 24
        sheet.column_dimensions["D"].width = 28

        sheet.merge_cells("A1:D1")
        sheet["A1"] = "Analysis Result"
        sheet["A1"].font = TITLE_FONT
        sheet["A1"].alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 30

        row = 3
        export_metadata = {
            "Exported at": datetime.now().astimezone().replace(microsecond=0),
            **metadata,
        }
        if export_metadata:
            row = self._section_title(sheet, row, "Report information", 4)
            metadata_rows = list(export_metadata.items())
            for index in range(0, len(metadata_rows), 2):
                left = metadata_rows[index]
                sheet.cell(row, 1, self._excel_value(left[0])).font = LABEL_FONT
                sheet.cell(row, 2, self._excel_value(left[1]))
                if index + 1 < len(metadata_rows):
                    right = metadata_rows[index + 1]
                    sheet.cell(row, 3, self._excel_value(right[0])).font = LABEL_FONT
                    sheet.cell(row, 4, self._excel_value(right[1]))
                row += 1
            row += 1

        if result.summary:
            row = self._section_title(sheet, row, "Analysis summary", 4)
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            cell = sheet.cell(row, 1, self._excel_value(result.summary))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            sheet.row_dimensions[row].height = min(
                120,
                max(36, 15 * (str(result.summary).count("\n") + 2)),
            )
            row += 2

        if result.metrics:
            row = self._section_title(sheet, row, "Key metrics", 4)
            self._write_header(sheet, row, ["Metric", "Value", "Unit", "Detail"])
            row += 1
            for metric in result.metrics:
                values = [metric.label, metric.value, metric.unit, metric.detail]
                for column, value in enumerate(values, start=1):
                    cell = sheet.cell(row, column, self._excel_value(value))
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                row += 1
            row += 1

        if result.insights:
            row = self._section_title(sheet, row, "Findings", 4)
            self._write_header(sheet, row, ["Type", "Title", "Detail", ""])
            row += 1
            for insight in result.insights:
                values = [insight.kind.title(), insight.title, insight.detail]
                for column, value in enumerate(values, start=1):
                    cell = sheet.cell(row, column, self._excel_value(value))
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if insight.kind == "warning":
                        cell.fill = WARNING_FILL
                row += 1
            row += 1

        if result.completed_requirements:
            row = self._section_title(sheet, row, "Completed requirements", 4)
            for requirement in result.completed_requirements:
                sheet.merge_cells(
                    start_row=row,
                    start_column=1,
                    end_row=row,
                    end_column=4,
                )
                cell = sheet.cell(row, 1, f"• {self._safe_text(requirement)}")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                row += 1

        self._apply_body_style(sheet)

    def _write_table(self, sheet, table) -> None:
        if len(table.columns) > EXCEL_MAX_COLUMNS:
            raise ValueError(
                f"Table {table.title!r} exceeds Excel's column limit."
            )
        if len(table.rows) + 4 > EXCEL_MAX_ROWS:
            raise ValueError(f"Table {table.title!r} exceeds Excel's row limit.")

        sheet.sheet_view.showGridLines = False
        sheet["A1"] = self._safe_text(table.title or "Analysis table")
        sheet["A1"].font = TITLE_FONT
        status = (
            f"Exported {len(table.rows):,} of {table.total_rows:,} rows"
            if table.truncated
            else f"{table.total_rows:,} rows"
        )
        sheet["A2"] = status
        sheet["A2"].font = Font(color="5F6368", italic=table.truncated)

        headers = list(table.columns)
        width = max(len(headers), max((len(row) for row in table.rows), default=0))
        if width == 0:
            sheet["A4"] = "No tabular data was returned."
            return
        if len(headers) < width:
            headers.extend(
                f"Column {index + 1}" for index in range(len(headers), width)
            )
        self._write_header(sheet, 4, headers)
        for row_index, values in enumerate(table.rows, start=5):
            padded = list(values[:width]) + [None] * max(0, width - len(values))
            for column_index, value in enumerate(padded, start=1):
                cell = sheet.cell(
                    row_index,
                    column_index,
                    self._excel_value(value),
                )
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        last_row = max(4, 4 + len(table.rows))
        sheet.auto_filter.ref = f"A4:{get_column_letter(width)}{last_row}"
        sheet.freeze_panes = "A5"
        self._size_columns(sheet, headers, table.rows)
        self._apply_body_style(sheet)

    def _write_charts(
        self,
        sheet,
        result: AnalysisResult,
        cancel_callback: Callable[[], bool] | None,
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 82
        sheet["A1"] = "Charts"
        sheet["A1"].font = TITLE_FONT
        row = 3
        for index, chart in enumerate(result.charts, start=1):
            self._check_cancel(cancel_callback)
            sheet.merge_cells(
                start_row=row,
                start_column=1,
                end_row=row,
                end_column=2,
            )
            sheet.cell(row, 1, self._safe_text(chart.title or f"Chart {index}"))
            sheet.cell(row, 1).font = SECTION_FONT
            row += 1
            try:
                image_bytes = base64.b64decode(
                    chart.image_base64,
                    validate=True,
                )
                image = WorksheetImage(io.BytesIO(image_bytes))
                scale = min(1.0, 760 / max(1, image.width), 420 / max(1, image.height))
                image.width = max(1, int(image.width * scale))
                image.height = max(1, int(image.height * scale))
                sheet.add_image(image, f"A{row}")
                image_rows = max(12, math.ceil(image.height / 20))
                row += image_rows
            except Exception:
                logger.warning(
                    "Chart image could not be exported title=%s",
                    chart.title,
                    exc_info=True,
                )
                sheet.cell(row, 1, "Chart image is unavailable.")
                row += 2
            if chart.caption:
                sheet.merge_cells(
                    start_row=row,
                    start_column=1,
                    end_row=row,
                    end_column=2,
                )
                caption = sheet.cell(row, 1, self._safe_text(chart.caption))
                caption.font = Font(color="5F6368", italic=True)
                caption.alignment = Alignment(wrap_text=True, vertical="top")
                row += 2
            else:
                row += 1

    def _write_audit(self, sheet, records: list[dict[str, Any]]) -> None:
        sheet.sheet_view.showGridLines = False
        columns = list(
            dict.fromkeys(str(key) for record in records for key in record)
        )
        if not columns:
            sheet["A1"] = "No audit records were returned."
            return
        self._write_header(sheet, 1, columns)
        for row_index, record in enumerate(records, start=2):
            for column_index, column in enumerate(columns, start=1):
                cell = sheet.cell(
                    row_index,
                    column_index,
                    self._excel_value(record.get(column)),
                )
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(columns))}{len(records) + 1}"
        )
        sheet.freeze_panes = "A2"
        self._size_columns(
            sheet,
            columns,
            [[record.get(column) for column in columns] for record in records],
        )
        self._apply_body_style(sheet)

    @staticmethod
    def _section_title(sheet, row: int, title: str, width: int) -> int:
        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=width,
        )
        cell = sheet.cell(row, 1, title)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[row].height = 23
        return row + 1

    @staticmethod
    def _write_header(sheet, row: int, headers: list[Any]) -> None:
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row, column, AnalysisExportService._excel_value(header))
            cell.fill = HEADER_FILL
            cell.font = WHITE_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = Border(bottom=THIN_GRAY)
        sheet.row_dimensions[row].height = 24

    @staticmethod
    def _apply_body_style(sheet) -> None:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None or cell.font == WHITE_FONT:
                    continue
                if not cell.font.bold:
                    cell.font = BODY_FONT

    @staticmethod
    def _size_columns(sheet, headers: list[Any], rows: list[list[Any]]) -> None:
        for index, header in enumerate(headers, start=1):
            sample_lengths = [len(str(header or ""))]
            for row in rows[:100]:
                if index - 1 < len(row):
                    sample_lengths.append(
                        max(
                            (len(line) for line in str(row[index - 1] or "").splitlines()),
                            default=0,
                        )
                    )
            sheet.column_dimensions[get_column_letter(index)].width = min(
                42,
                max(10, max(sample_lengths, default=10) + 2),
            )

    @staticmethod
    def _unique_sheet_name(name: str, used_names: set[str]) -> str:
        base = re.sub(r"[\[\]:*?/\\]", "_", str(name)).strip()[:31] or "Sheet"
        candidate = base
        index = 2
        lowered = {item.lower() for item in used_names}
        while candidate.lower() in lowered:
            suffix = f" ({index})"
            candidate = f"{base[: 31 - len(suffix)]}{suffix}"
            index += 1
        return candidate

    @staticmethod
    def _safe_text(value: Any) -> str:
        text = str(value or "")
        if len(text) > EXCEL_MAX_CELL_CHARACTERS:
            text = text[: EXCEL_MAX_CELL_CHARACTERS - 1] + "…"
        if text.startswith(FORMULA_PREFIXES):
            return "'" + text
        return text

    @classmethod
    def _excel_value(cls, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return value if math.isfinite(value) else None
        if isinstance(value, datetime):
            if value.tzinfo is not None:
                return value.astimezone().replace(tzinfo=None)
            return value
        if isinstance(value, date):
            return value
        if hasattr(value, "item"):
            try:
                return cls._excel_value(value.item())
            except (TypeError, ValueError):
                pass
        return cls._safe_text(value)

    @staticmethod
    def _check_cancel(callback: Callable[[], bool] | None) -> None:
        if callback and callback():
            raise RuntimeError("Export cancelled.")
