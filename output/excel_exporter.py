"""
output/excel_exporter.py
────────────────────────
将分析结果导出为 Excel 文件。

Sprint 3 🟡
"""

from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from output.formatter import FormattedOutput


class ExcelExporter:
    """Excel 导出器"""

    def export(
        self,
        result: FormattedOutput,
        output_path: str,
        title: str = "数据分析结果",
    ) -> None:
        """
        将分析结果导出为 Excel。
        
        Args:
            result: FormattedOutput 对象
            output_path: 输出文件路径
            title: 工作表标题
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "分析结果"

        # 标题行
        ws["A1"] = title
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:D1")

        row = 3

        # 结果内容
        if result.has_error:
            ws[f"A{row}"] = "❌ 错误信息"
            ws[f"A{row}"].font = Font(color="FF0000", bold=True)
            row += 1
            ws[f"A{row}"] = result.error
            row += 2
        else:
            # 输出文本
            ws[f"A{row}"] = "📊 分析结果"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            for line in result.text_blocks:
                ws[f"A{row}"] = line
                row += 1

            row += 1

            # 表格数据
            if result.has_table:
                ws[f"A{row}"] = "📋 数据表"
                ws[f"A{row}"].font = Font(bold=True)
                row += 1

                headers = list(result.table_data[0].keys())
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                row += 1

                for record in result.table_data:
                    for col, header in enumerate(headers, start=1):
                        ws.cell(row=row, column=col, value=record.get(header, ""))
                    row += 1

                row += 1

        # 元数据
        ws[f"A{row}"] = "ℹ️ 执行信息"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        ws[f"A{row}"] = f"耗时: {result.elapsed_sec}s"
        row += 1
        ws[f"A{row}"] = f"重试次数: {result.retries_used}"
        row += 1

        # 代码
        if result.code:
            row += 1
            ws[f"A{row}"] = "💻 执行的代码"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            code_sheet = wb.create_sheet("代码")
            code_sheet["A1"] = result.code
            code_sheet["A1"].alignment = Alignment(wrap_text=True)

        # 调整列宽
        ws.column_dimensions["A"].width = 50

        # 保存
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)