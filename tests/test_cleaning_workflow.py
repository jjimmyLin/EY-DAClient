from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import load_workbook

from config.settings import settings
from core.cleaning_service import CleaningIssue, CleaningProfile, CleaningService
from core.preprocessor import FileMeta, Preprocessor, ProcessingCancelled, SheetMeta
from ui.main_window import MainWindow


def _prepare_dataset(tmp_path, monkeypatch):
    cache_dir = tmp_path / "cache"
    temp_dir = tmp_path / "duckdb"
    monkeypatch.setattr(settings, "DATASET_CACHE_DIR", cache_dir)
    monkeypatch.setattr(settings, "DUCKDB_TEMP_DIR", temp_dir)
    monkeypatch.setattr(settings, "IMPORT_MIN_FREE_DISK_BYTES", 0)
    source = tmp_path / "dirty.xlsx"
    pd.DataFrame(
        {
            "name": [" Alice ", " Alice ", None, "Bob"],
            "amount": [10.0, 10.0, None, 30.0],
            "empty": [None, None, None, None],
        }
    ).to_excel(source, index=False)
    return Preprocessor().process(str(source))


def _parquet_file_meta(tmp_path, dataframe: pd.DataFrame) -> FileMeta:
    cache = tmp_path / "source.parquet"
    dataframe.to_parquet(cache, index=False)
    sheet = SheetMeta(
        sheet_name="Sheet1",
        rows=len(dataframe),
        cols=len(dataframe.columns),
        columns=[str(column) for column in dataframe.columns],
        dtypes={
            str(column): str(dtype)
            for column, dtype in dataframe.dtypes.items()
        },
        null_counts={
            str(column): int(count)
            for column, count in dataframe.isna().sum().items()
            if count
        },
        head_sample=[],
        describe={},
        unique_values={},
        cache_path=str(cache),
    )
    return FileMeta(
        file_path=str(tmp_path / "original.xlsx"),
        file_name="original.xlsx",
        file_size_kb=max(1, cache.stat().st_size / 1024),
        sheet_count=1,
        sheets=[sheet],
        dataset_id="ds_test",
    )


def test_cleaning_profile_detects_supported_dirty_data(tmp_path, monkeypatch):
    file_meta = _prepare_dataset(tmp_path, monkeypatch)

    profile = CleaningService().profile(file_meta)
    issue_ids = {issue.issue_id for issue in profile.issues}

    assert "missing_values" in issue_ids
    assert "duplicate_rows" in issue_ids
    assert "key_duplicates" in issue_ids
    assert "blank_rows" in issue_ids
    assert "empty_columns" in issue_ids


def test_cleaning_executes_selected_rules_and_preserves_source(tmp_path, monkeypatch):
    file_meta = _prepare_dataset(tmp_path, monkeypatch)
    output = tmp_path / "cleaned.xlsx"

    result = CleaningService().clean(
        file_meta,
        {
            "missing_values": "fill_zero",
            "duplicate_rows": "duplicate_keep_first",
            "blank_rows": "drop_blank_rows",
            "empty_columns": "drop_empty_columns",
        },
        str(output),
    )

    cleaned = pd.read_excel(output)
    assert list(cleaned.columns) == ["name", "amount"]
    assert cleaned["name"].tolist() == [" Alice ", "Bob"]
    assert cleaned["amount"].tolist() == [10, 30]
    assert result.rows_before == 4
    assert result.rows_after == 2
    assert file_meta.file_path.endswith("dirty.xlsx")


def test_cleaning_page_is_available_from_navigation(qapp):
    window = MainWindow()
    window._start_new_task()

    assert window.nav_clean_btn.isEnabled()
    window._show_cleaning_page()
    assert window.page_container.currentWidget() is window.cleaning_page

    window.close()


def test_cleaning_rules_are_always_visible_and_start_disabled(qapp):
    window = MainWindow()
    page = window.cleaning_page

    assert set(page._rule_cards) == {
        "missing_values",
        "duplicate_rows",
        "key_duplicates",
        "blank_rows",
        "empty_columns",
        "mixed_numeric_values",
    }
    assert all(
        not checkbox.isEnabled()
        for checkbox, _ in page._issue_rows.values()
    )
    assert sum(
        len(selector._buttons)
        for _, selector in page._issue_rows.values()
    ) == 20
    assert all(
        not any(button.isChecked() for button in selector._buttons.values())
        for _, selector in page._issue_rows.values()
    )

    window.close()


def test_duplicate_rules_support_keep_last_and_remove_all(tmp_path, monkeypatch):
    cache_dir = tmp_path / "cache"
    temp_dir = tmp_path / "duckdb"
    monkeypatch.setattr(settings, "DATASET_CACHE_DIR", cache_dir)
    monkeypatch.setattr(settings, "DUCKDB_TEMP_DIR", temp_dir)
    monkeypatch.setattr(settings, "IMPORT_MIN_FREE_DISK_BYTES", 0)
    source = tmp_path / "duplicates.xlsx"
    pd.DataFrame(
        {
            "id": [1, 1, 2],
            "value": ["same", "same", "unique"],
        }
    ).to_excel(source, index=False)
    file_meta = Preprocessor().process(str(source))

    keep_last = tmp_path / "keep-last.xlsx"
    CleaningService().clean(
        file_meta,
        {"duplicate_rows": "duplicate_keep_last"},
        str(keep_last),
    )
    assert pd.read_excel(keep_last).to_dict("records") == [
        {"id": 1, "value": "same"},
        {"id": 2, "value": "unique"},
    ]

    remove_all = tmp_path / "remove-all.xlsx"
    CleaningService().clean(
        file_meta,
        {"duplicate_rows": "duplicate_remove_all"},
        str(remove_all),
    )
    assert pd.read_excel(remove_all).to_dict("records") == [
        {"id": 2, "value": "unique"},
    ]


def test_key_duplicate_rule_uses_user_selected_composite_key(tmp_path, monkeypatch):
    cache_dir = tmp_path / "cache"
    temp_dir = tmp_path / "duckdb"
    monkeypatch.setattr(settings, "DATASET_CACHE_DIR", cache_dir)
    monkeypatch.setattr(settings, "DUCKDB_TEMP_DIR", temp_dir)
    monkeypatch.setattr(settings, "IMPORT_MIN_FREE_DISK_BYTES", 0)
    source = tmp_path / "key-duplicates.xlsx"
    pd.DataFrame(
        {
            "order": ["A", "A", "A", "B"],
            "line": [1, 1, 2, 1],
            "amount": [10, 20, 30, 40],
        }
    ).to_excel(source, index=False)
    file_meta = Preprocessor().process(str(source))
    sheet_name = file_meta.sheets[0].sheet_name
    output = tmp_path / "key-cleaned.xlsx"

    CleaningService().clean(
        file_meta,
        {
            "key_duplicates": {
                "method": "key_keep_last",
                "columns": {sheet_name: ["order", "line"]},
            }
        },
        str(output),
    )

    assert pd.read_excel(output).to_dict("records") == [
        {"order": "A", "line": 1, "amount": 20},
        {"order": "A", "line": 2, "amount": 30},
        {"order": "B", "line": 1, "amount": 40},
    ]


def test_full_and_key_duplicate_rules_are_applied_sequentially(
    tmp_path,
    monkeypatch,
):
    cache_dir = tmp_path / "cache"
    temp_dir = tmp_path / "duckdb"
    monkeypatch.setattr(settings, "DATASET_CACHE_DIR", cache_dir)
    monkeypatch.setattr(settings, "DUCKDB_TEMP_DIR", temp_dir)
    monkeypatch.setattr(settings, "IMPORT_MIN_FREE_DISK_BYTES", 0)
    source = tmp_path / "combined-duplicates.xlsx"
    pd.DataFrame(
        {
            "id": ["A", "A", "A", "B"],
            "amount": [10, 10, 20, 30],
        }
    ).to_excel(source, index=False)
    file_meta = Preprocessor().process(str(source))
    sheet_name = file_meta.sheets[0].sheet_name
    output = tmp_path / "combined-cleaned.xlsx"

    CleaningService().clean(
        file_meta,
        {
            "duplicate_rows": "duplicate_keep_first",
            "key_duplicates": {
                "method": "key_keep_last",
                "columns": {sheet_name: ["id"]},
            },
        },
        str(output),
    )

    assert pd.read_excel(output).to_dict("records") == [
        {"id": "A", "amount": 20},
        {"id": "B", "amount": 30},
    ]


def test_key_duplicate_rule_preserves_rows_with_missing_keys(
    tmp_path,
    monkeypatch,
):
    cache_dir = tmp_path / "cache"
    temp_dir = tmp_path / "duckdb"
    monkeypatch.setattr(settings, "DATASET_CACHE_DIR", cache_dir)
    monkeypatch.setattr(settings, "DUCKDB_TEMP_DIR", temp_dir)
    monkeypatch.setattr(settings, "IMPORT_MIN_FREE_DISK_BYTES", 0)
    source = tmp_path / "missing-keys.xlsx"
    pd.DataFrame(
        {
            "id": [None, None, "A", "A"],
            "amount": [10, 20, 30, 40],
        }
    ).to_excel(source, index=False)
    file_meta = Preprocessor().process(str(source))
    sheet_name = file_meta.sheets[0].sheet_name
    output = tmp_path / "missing-keys-cleaned.xlsx"

    CleaningService().clean(
        file_meta,
        {
            "key_duplicates": {
                "method": "key_keep_first",
                "columns": {sheet_name: ["id"]},
            }
        },
        str(output),
    )

    cleaned = pd.read_excel(output)
    assert cleaned["amount"].tolist() == [10, 20, 30]


def test_missing_row_deletion_ignores_columns_that_are_entirely_empty(
    tmp_path,
    monkeypatch,
):
    file_meta = _prepare_dataset(tmp_path, monkeypatch)
    output = tmp_path / "drop-missing.xlsx"

    CleaningService().clean(
        file_meta,
        {"missing_values": "drop_rows"},
        str(output),
    )

    cleaned = pd.read_excel(output)
    assert len(cleaned) == 3
    assert cleaned["name"].tolist() == [" Alice ", " Alice ", "Bob"]


def test_rule_selection_does_not_resize_scroll_content(qapp):
    window = MainWindow()
    page = window.cleaning_page
    profile = CleaningProfile(
        dataset_id="ds_test",
        rows=10,
        columns=4,
        sheets=1,
        issues=[
            CleaningIssue(
                issue_id,
                card.definition["title"],
                2,
                methods=list(card.definition["methods"]),
                sheet_columns={"Sheet1": ["id", "value"]}
                if issue_id == "key_duplicates"
                else {},
            )
            for issue_id, card in page._rule_cards.items()
        ],
    )
    page.set_target_dataset("test.xlsx")
    page.show_profile(profile)
    qapp.processEvents()
    initial_height = page.cards_container.sizeHint().height()

    for card in page._rule_cards.values():
        card.checkbox.setChecked(True)
    qapp.processEvents()

    assert page.cards_container.sizeHint().height() == initial_height
    assert all(
        not card.treatment_widget.isHidden()
        for card in page._rule_cards.values()
    )
    window.close()


def test_column_issue_ui_shows_counts_and_emits_column_configuration(qapp):
    window = MainWindow()
    page = window.cleaning_page
    issue = CleaningIssue(
        "missing_values",
        "Missing values",
        5,
        methods=["drop_rows", "fill_zero", "fill_mode"],
        column_counts={"Sheet1": {"amount": 3, "label": 2}},
        column_methods={
            "Sheet1": {
                "amount": ["drop_rows", "fill_zero"],
                "label": ["drop_rows", "fill_mode"],
            }
        },
        estimated_changes=5,
    )
    page.set_target_dataset("test.xlsx")
    page.show_profile(
        CleaningProfile("ds_test", 10, 2, 1, [issue])
    )
    card = page._rule_cards["missing_values"]
    card.checkbox.setChecked(True)
    card._selected_column_methods = {
        "Sheet1": {"amount": "fill_zero", "label": "fill_mode"}
    }
    page._refresh_execute_state()

    assert "Sheet1.amount" in card.affected_label.text()
    assert "Sheet1.label" in card.affected_label.text()
    assert "5 affected value(s)" in card.affected_label.text()
    assert card.selection_payload() == {
        "columns": {
            "Sheet1": {"amount": "fill_zero", "label": "fill_mode"}
        }
    }
    assert page.execute_button.isEnabled()
    window.close()


def test_cleaning_rule_sections_have_compact_buttons_and_clear_states(qapp):
    window = MainWindow()
    window.show()
    window._show_cleaning_page()
    page = window.cleaning_page
    page.set_target_dataset("test.xlsx")
    page.show_profile(
        CleaningProfile(
            "ds_test",
            10,
            2,
            1,
            [
                CleaningIssue(
                    "duplicate_rows",
                    "Duplicate rows",
                    2,
                    methods=[
                        "duplicate_keep_first",
                        "duplicate_keep_last",
                        "duplicate_remove_all",
                    ],
                    estimated_changes=2,
                )
            ],
        )
    )
    qapp.processEvents()

    available = page._rule_cards["duplicate_rows"]
    unavailable = page._rule_cards["blank_rows"]
    assert available.property("ruleState") == "available"
    assert unavailable.property("ruleState") == "unavailable"
    assert all(
        button.height() == 25 and button.width() < 180
        for button in available.options._buttons.values()
    )

    available.checkbox.setChecked(True)
    qapp.processEvents()
    assert available.property("selected") is True
    assert any(
        button.isChecked()
        for button in available.options._buttons.values()
    )
    window.close()


def test_stale_cleaning_profile_is_not_rendered_for_new_target(qapp):
    window = MainWindow()
    window.cleaning_page.set_target_dataset("new.xlsx")
    original_status = window.cleaning_page.status_label.text()
    stale_profile = CleaningProfile(
        "ds_old",
        10,
        2,
        1,
        [
            CleaningIssue(
                "missing_values",
                "Missing values",
                2,
                methods=["drop_rows"],
            )
        ],
    )

    window._on_cleaning_profile_finished("old.xlsx", stale_profile)

    assert window.cleaning_page.status_label.text() == original_status
    assert "Ignored stale cleaning scan" in window.log_output.toPlainText()
    window.close()


def test_cleaning_service_cancel_stops_before_work(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "DUCKDB_TEMP_DIR", tmp_path / "duckdb")
    file_meta = _parquet_file_meta(
        tmp_path,
        pd.DataFrame({"value": [1, 2]}),
    )
    service = CleaningService()
    service.cancel()

    with pytest.raises(ProcessingCancelled):
        service.profile(file_meta)


def test_cleaning_profile_reports_background_progress(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "DUCKDB_TEMP_DIR", tmp_path / "duckdb")
    file_meta = _parquet_file_meta(
        tmp_path,
        pd.DataFrame({"value": [1, 1, 2]}),
    )
    progress = []

    CleaningService().profile(
        file_meta,
        progress_callback=lambda percent, detail: progress.append(
            (percent, detail)
        ),
    )

    assert progress[0][0] == 0
    assert progress[-1] == (100, "Scan complete")


def test_cleaning_cancel_button_cancels_active_worker(qapp):
    window = MainWindow()
    cancelled = []
    window._cleaning_worker = type(
        "Worker",
        (),
        {"cancel": lambda self: cancelled.append(True)},
    )()
    window.cleaning_page.cancel_button.setVisible(True)

    window._cancel_cleaning()

    assert cancelled == [True]
    assert not window.cleaning_page.cancel_button.isEnabled()
    assert "cancellation requested" in window.log_output.toPlainText()
    window._cleaning_worker = None
    window.close()


def test_mixed_numeric_column_imports_without_data_loss(tmp_path, monkeypatch):
    cache_dir = tmp_path / "cache"
    temp_dir = tmp_path / "duckdb"
    monkeypatch.setattr(settings, "DATASET_CACHE_DIR", cache_dir)
    monkeypatch.setattr(settings, "DUCKDB_TEMP_DIR", temp_dir)
    monkeypatch.setattr(settings, "IMPORT_MIN_FREE_DISK_BYTES", 0)
    source = tmp_path / "mixed-price.xlsx"
    pd.DataFrame(
        {"item": ["A", "B", "C", "D"], "price": [10, 20, "价格待定", 40]}
    ).to_excel(source, index=False)

    file_meta = Preprocessor().process(str(source))
    sheet = file_meta.sheets[0]
    cached = pd.read_parquet(sheet.cache_path)

    assert cached["price"].tolist() == ["10", "20", "价格待定", "40"]
    assert sheet.type_profiles["price"]["inferred_type"] == "numeric"
    assert sheet.type_profiles["price"]["invalid_count"] == 1


def test_cleaning_detects_and_repairs_invalid_numeric_values(tmp_path, monkeypatch):
    cache_dir = tmp_path / "cache"
    temp_dir = tmp_path / "duckdb"
    monkeypatch.setattr(settings, "DATASET_CACHE_DIR", cache_dir)
    monkeypatch.setattr(settings, "DUCKDB_TEMP_DIR", temp_dir)
    monkeypatch.setattr(settings, "IMPORT_MIN_FREE_DISK_BYTES", 0)
    source = tmp_path / "mixed-price.xlsx"
    pd.DataFrame({"price": [10, 20, "价格待定", 40]}).to_excel(source, index=False)
    file_meta = Preprocessor().process(str(source))

    profile = CleaningService().profile(file_meta)
    issue = next(
        issue
        for issue in profile.issues
        if issue.issue_id == "mixed_numeric_values"
    )
    assert issue.count == 1

    output = tmp_path / "cleaned-price.xlsx"
    CleaningService().clean(
        file_meta,
        {"mixed_numeric_values": "invalid_mean"},
        str(output),
    )
    cleaned = pd.read_excel(output)

    assert cleaned["price"].iloc[0] == 10
    assert cleaned["price"].iloc[1] == 20
    assert cleaned["price"].iloc[2] == pytest.approx(70 / 3)
    assert cleaned["price"].iloc[3] == 40


def test_column_level_missing_methods_are_applied_independently(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setattr(settings, "DUCKDB_TEMP_DIR", tmp_path / "duckdb")
    monkeypatch.setattr(settings, "CLEANING_MIN_FREE_DISK_BYTES", 0)
    file_meta = _parquet_file_meta(
        tmp_path,
        pd.DataFrame(
            {
                "amount": [10.0, None, 30.0],
                "quantity": [1.0, None, 3.0],
                "label": ["A", None, "C"],
            }
        ),
    )
    output = tmp_path / "column-cleaned.xlsx"

    CleaningService().clean(
        file_meta,
        {
            "missing_values": {
                "columns": {
                    "Sheet1": {
                        "amount": "fill_mean",
                        "quantity": "fill_zero",
                        "label": "fill_mode",
                    }
                }
            }
        },
        str(output),
    )

    cleaned = pd.read_excel(output)
    assert cleaned["amount"].tolist() == [10, 20, 30]
    assert cleaned["quantity"].tolist() == [1, 0, 3]
    assert cleaned["label"].tolist() == ["A", "A", "C"]


def test_invalid_numeric_replacement_preserves_true_nulls(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setattr(settings, "DUCKDB_TEMP_DIR", tmp_path / "duckdb")
    monkeypatch.setattr(settings, "CLEANING_MIN_FREE_DISK_BYTES", 0)
    dataframe = pd.DataFrame(
        {"price": pd.Series(["10", None, "bad", "20"], dtype="string")}
    )
    file_meta = _parquet_file_meta(tmp_path, dataframe)
    file_meta.sheets[0].type_profiles["price"] = {
        "inferred_type": "numeric",
        "invalid_count": 1,
    }
    output = tmp_path / "numeric-cleaned.xlsx"

    CleaningService().clean(
        file_meta,
        {
            "mixed_numeric_values": {
                "columns": {"Sheet1": {"price": "invalid_zero"}}
            }
        },
        str(output),
    )

    cleaned = pd.read_excel(output)
    assert cleaned["price"].iloc[0] == 10
    assert pd.isna(cleaned["price"].iloc[1])
    assert cleaned["price"].iloc[2] == 0
    assert cleaned["price"].iloc[3] == 20


def test_invalid_to_null_stays_null_when_original_nulls_are_filled(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setattr(settings, "DUCKDB_TEMP_DIR", tmp_path / "duckdb")
    monkeypatch.setattr(settings, "CLEANING_MIN_FREE_DISK_BYTES", 0)
    dataframe = pd.DataFrame(
        {"price": pd.Series(["10", None, "bad", "20"], dtype="string")}
    )
    file_meta = _parquet_file_meta(tmp_path, dataframe)
    file_meta.sheets[0].type_profiles["price"] = {
        "inferred_type": "numeric",
        "invalid_count": 1,
    }
    output = tmp_path / "separated-null.xlsx"

    CleaningService().clean(
        file_meta,
        {
            "missing_values": {
                "columns": {"Sheet1": {"price": "fill_mean"}}
            },
            "mixed_numeric_values": {
                "columns": {"Sheet1": {"price": "invalid_to_null"}}
            },
        },
        str(output),
    )

    cleaned = pd.read_excel(output)
    assert cleaned["price"].iloc[0] == 10
    assert cleaned["price"].iloc[1] == 15
    assert pd.isna(cleaned["price"].iloc[2])
    assert cleaned["price"].iloc[3] == 20


def test_cleaning_refuses_to_overwrite_original_file(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "DUCKDB_TEMP_DIR", tmp_path / "duckdb")
    monkeypatch.setattr(settings, "CLEANING_MIN_FREE_DISK_BYTES", 0)
    file_meta = _parquet_file_meta(
        tmp_path,
        pd.DataFrame({"value": [1, 1]}),
    )

    with pytest.raises(ValueError, match="cannot overwrite"):
        CleaningService().clean(
            file_meta,
            {"duplicate_rows": "duplicate_keep_first"},
            file_meta.file_path,
        )


def test_excel_output_escapes_formulas_and_illegal_characters(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setattr(settings, "DUCKDB_TEMP_DIR", tmp_path / "duckdb")
    monkeypatch.setattr(settings, "CLEANING_MIN_FREE_DISK_BYTES", 0)
    file_meta = _parquet_file_meta(
        tmp_path,
        pd.DataFrame({"text": ["=1+1", "@SUM(A1:A2)", "bad\x01text"]}),
    )
    output = tmp_path / "safe.xlsx"

    CleaningService().clean(
        file_meta,
        {"duplicate_rows": "duplicate_keep_first"},
        str(output),
    )

    workbook = load_workbook(output, data_only=False)
    values = [workbook["Sheet1"].cell(row, 1).value for row in range(2, 5)]
    assert values == ["'=1+1", "'@SUM(A1:A2)", "badtext"]
    assert all(
        workbook["Sheet1"].cell(row, 1).data_type != "f"
        for row in range(2, 5)
    )


def test_cleaning_disk_guard_fails_before_writing(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "DUCKDB_TEMP_DIR", tmp_path / "duckdb")
    monkeypatch.setattr(settings, "CLEANING_MIN_FREE_DISK_BYTES", 1024)
    monkeypatch.setattr(
        "core.cleaning_service.shutil.disk_usage",
        lambda path: type("Usage", (), {"free": 1})(),
    )
    file_meta = _parquet_file_meta(
        tmp_path,
        pd.DataFrame({"value": [1, 1]}),
    )
    output = tmp_path / "no-space.xlsx"

    with pytest.raises(OSError, match="Insufficient disk space"):
        CleaningService().clean(
            file_meta,
            {"duplicate_rows": "duplicate_keep_first"},
            str(output),
        )

    assert not output.exists()


def test_new_analysis_preserves_shared_dataset_library(qapp):
    window = MainWindow()
    file_meta = FileMeta(
        file_path="C:/shared.xlsx",
        file_name="shared.xlsx",
        file_size_kb=10,
        sheet_count=0,
        sheets=[],
        dataset_id="ds_shared",
    )
    window.loaded_files["shared.xlsx"] = file_meta
    window._dataset_states["shared.xlsx"] = {
        "state": "ready",
        "file_path": file_meta.file_path,
        "percent": 100,
    }
    window._add_dataset_item("shared.xlsx")

    window._start_new_task()
    window.prompt_input.setPlainText("temporary question")
    window._start_new_task()

    assert window.loaded_files["shared.xlsx"] is file_meta
    assert window.dataset_list.count() == 1
    assert window.prompt_input.toPlainText() == ""

    window.close()


def test_cleaning_uses_library_independently_of_analysis_scope(qapp):
    window = MainWindow()
    file_meta = FileMeta(
        file_path="C:/clean-me.xlsx",
        file_name="clean-me.xlsx",
        file_size_kb=10,
        sheet_count=0,
        sheets=[],
        dataset_id="ds_clean",
    )
    window.loaded_files["clean-me.xlsx"] = file_meta
    window._add_dataset_item("clean-me.xlsx", selected=False)

    window._start_cleaning()

    assert window.cleaning_page.target_label.text() == "clean-me.xlsx"
    assert window._selected_datasets == {"clean-me.xlsx"}
    assert window.loaded_files["clean-me.xlsx"] is file_meta

    window.close()
