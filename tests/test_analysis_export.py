from __future__ import annotations

from openpyxl import load_workbook

from core.analysis_export import AnalysisExportService
from core.analysis_result import (
    AnalysisResult,
    AnswerResult,
    InsightResult,
    MetricResult,
    TableResult,
)


def test_analysis_result_exports_to_structured_excel(tmp_path):
    destination = tmp_path / "analysis-result.xlsx"
    result = AnalysisResult(
        summary="Product A leads total revenue.",
        answers=[
            AnswerResult(
                "A",
                "Calculate revenue by product",
                "Product A leads with 80 CNY.",
                supporting_metrics=["Total revenue"],
                supporting_tables=["Revenue by product"],
                confidence_or_notes="Calculated from price and quantity.",
            )
        ],
        metrics=[
            MetricResult("Total revenue", 115.0, "CNY", "Across all products"),
        ],
        tables=[
            TableResult(
                "Revenue by product",
                ["Product", "Revenue"],
                [["A", 80], ["B", 35]],
                total_rows=2,
            )
        ],
        insights=[
            InsightResult(
                "Leading product",
                "Product A contributes the largest share.",
            ),
            InsightResult(
                "Data quality",
                "Three rows have missing prices.",
                "warning",
            ),
        ],
        audit=[{"step": "aggregate", "rows": 2}],
        completed_requirements=["A", "B"],
    )

    output = AnalysisExportService().export_excel(
        result,
        destination,
        metadata={"Datasets": "sales.xlsx", "Request": "Review revenue"},
    )

    assert output == str(destination)
    assert destination.exists()
    assert not (tmp_path / "analysis-result.xlsx.partial").exists()

    workbook = load_workbook(destination, data_only=False)
    assert workbook.sheetnames == ["Summary", "Revenue by product", "Audit"]
    summary = workbook["Summary"]
    assert summary["A1"].value == "Analysis Result"
    assert any(
        cell.value == "Product A leads total revenue."
        for row in summary.iter_rows()
        for cell in row
    )
    assert any(
        cell.value == "Calculate revenue by product"
        for row in summary.iter_rows()
        for cell in row
    )
    assert any(
        cell.value == "Product A leads with 80 CNY."
        for row in summary.iter_rows()
        for cell in row
    )
    table = workbook["Revenue by product"]
    assert table["A4"].value == "Product"
    assert table["B5"].value == 80
    assert table.auto_filter.ref == "A4:B6"
    assert workbook["Audit"]["A2"].value == "aggregate"


def test_export_escapes_formula_like_text_and_deduplicates_sheet_names(tmp_path):
    destination = tmp_path / "safe.xlsx"
    result = AnalysisResult(
        summary="=HYPERLINK(\"https://example.invalid\",\"click\")",
        tables=[
            TableResult("Duplicate/name", ["Value"], [["=1+1"]], 1),
            TableResult("Duplicate:name", ["Value"], [["@SUM(A1:A2)"]], 1),
        ],
    )

    AnalysisExportService().export_excel(result, destination)

    workbook = load_workbook(destination, data_only=False)
    assert workbook["Summary"]["A6"].data_type != "f"
    assert workbook.sheetnames == [
        "Summary",
        "Duplicate_name",
        "Duplicate_name (2)",
    ]
    assert workbook["Duplicate_name"]["A5"].data_type != "f"
    assert workbook["Duplicate_name (2)"]["A5"].data_type != "f"


def test_export_removes_partial_file_after_failure(tmp_path, monkeypatch):
    destination = tmp_path / "broken.xlsx"

    def fail_save(self, path):
        raise OSError("disk full")

    monkeypatch.setattr("core.analysis_export.Workbook.save", fail_save)

    try:
        AnalysisExportService().export_excel(
            AnalysisResult(summary="Ready"),
            destination,
        )
    except OSError as exc:
        assert "disk full" in str(exc)
    else:
        raise AssertionError("export should have failed")

    assert not destination.exists()
    assert not (tmp_path / "broken.xlsx.partial").exists()
